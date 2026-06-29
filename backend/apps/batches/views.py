import csv
import io
from datetime import date

import openpyxl
from django.http import HttpResponse
from rest_framework import viewsets, status, filters, serializers
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.db import transaction
from django.db.models import Q

from apps.centers.models import Center
from apps.trainers.models import Trainer, TrainerMapping
from apps.trainees.models import Trainee
from apps.trainees.serializers import TraineeListSerializer
from .models import Batch, BatchWeekPlan, BatchEnrollment
from .serializers import (
    BatchListSerializer,
    BatchDetailSerializer,
    BatchWriteSerializer,
    BatchWeekPlanListSerializer,
    BatchWeekPlanWriteSerializer,
    BatchEnrollmentSerializer,
    BatchEnrollmentBulkSerializer,
    BatchCalendarDaySerializer,
    ShiftSerializer,
    HolidaySerializer,
)
from .validators import validate_batch_hours_match_course


class BatchViewSet(viewsets.ModelViewSet):
    queryset = Batch.objects.select_related(
        'circular', 'center', 'course', 'created_by'
    ).prefetch_related('week_plans', 'enrollments').all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ('center', 'course', 'circular', 'status')
    search_fields = ('batch_no', 'custom_batch_no', 'batch_name_bn', 'batch_name_en')
    ordering_fields = ('start_date', 'created_at', 'batch_name_bn')
    ordering = ('-start_date',)

    def get_serializer_class(self):
        if self.action == 'list':
            return BatchListSerializer
        elif self.action in ('create', 'update', 'partial_update'):
            return BatchWriteSerializer
        return BatchDetailSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.user_type == 'center_admin' and user.center:
            qs = qs.filter(center=user.center)
        elif user.user_type == 'center_admin' and not user.center:
            return Batch.objects.none()
        elif user.user_type == 'trainer':
            try:
                trainer = user.trainer_profile
            except Trainer.DoesNotExist:
                return Batch.objects.none()
            batch_ids = BatchWeekPlan.objects.filter(
                Q(lead_trainer=trainer) | Q(associate_trainer=trainer)
            ).values_list('batch_id', flat=True).distinct()
            qs = qs.filter(id__in=batch_ids)
        return qs

    def perform_create(self, serializer):
        self.assert_admin_or_ho(self.request)
        user = self.request.user
        data = {}
        if user.user_type == 'center_admin' and user.center:
            data['center'] = user.center
        data['created_by'] = user
        serializer.save(**data)

    def perform_update(self, serializer):
        self.assert_admin_or_ho(self.request)
        serializer.save()

    def perform_destroy(self, instance):
        self.assert_admin_or_ho(self.request)
        instance.delete()

    @action(detail=False, methods=['get'], permission_classes=[IsAuthenticated])
    def my_batches(self, request):
        if request.user.user_type != 'trainer':
            return Response({'detail': 'Only trainers can access this.'}, status=403)
        try:
            trainer = request.user.trainer_profile
        except Trainer.DoesNotExist:
            return Response({'detail': 'Trainer profile not found.'}, status=404)
        batch_ids = BatchWeekPlan.objects.filter(
            Q(lead_trainer=trainer) | Q(associate_trainer=trainer)
        ).values_list('batch_id', flat=True).distinct()
        batches = Batch.objects.filter(id__in=batch_ids).select_related(
            'circular', 'center', 'course'
        )
        page = self.paginate_queryset(batches)
        if page is not None:
            serializer = BatchListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = BatchListSerializer(batches, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def week_plans(self, request, pk=None):
        batch = self.get_object()
        plans = BatchWeekPlan.objects.filter(batch=batch).select_related(
            'lead_trainer__user', 'associate_trainer__user'
        )
        page = self.paginate_queryset(plans)
        if page is not None:
            serializer = BatchWeekPlanListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = BatchWeekPlanListSerializer(plans, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def assign_trainer(self, request, pk=None):
        batch = self.get_object()
        center_id = None
        if request.user.user_type == 'center_admin':
            center_id = request.user.center_id
        elif request.user.user_type != 'head_office':
            return Response({'error': 'অনুমতি নেই।'}, status=403)
        if center_id and str(batch.center_id) != str(center_id):
            return Response({'error': 'আপনি শুধুমাত্র নিজ কেন্দ্রের ব্যাচে নিয়োগ দিতে পারেন।'}, status=403)

        trainer_id = request.data.get('trainer_id')
        associate_id = request.data.get('associate_trainer_id')
        if not trainer_id:
            return Response({'error': 'প্রধান প্রশিক্ষক নির্বাচন করুন।'}, status=400)

        try:
            lead = Trainer.objects.select_related('user').get(
                pk=trainer_id, status='active', approval_status='approved',
            )
        except Trainer.DoesNotExist:
            return Response({'error': 'প্রধান প্রশিক্ষক খুঁজে পাওয়া যায়নি বা সক্রিয় নন।'}, status=400)

        if associate_id and str(associate_id) != str(trainer_id):
            try:
                associate = Trainer.objects.select_related('user').get(
                    pk=associate_id, status='active', approval_status='approved',
                )
            except Trainer.DoesNotExist:
                return Response({'error': 'সহকারী প্রশিক্ষক খুঁজে পাওয়া যায়নি বা সক্রিয় নন।'}, status=400)
        else:
            associate = None

        mapping_exists = TrainerMapping.objects.filter(
            trainer=lead, center=batch.center, course=batch.course, status='active',
        ).exists()
        if not mapping_exists and not request.user.is_superuser:
            err = 'প্রধান প্রশিক্ষক এই ব্যাচের কেন্দ্র ও কোর্সের জন্য নির্ধারিত নন।'
            return Response({'error': err}, status=400)

        if associate:
            assoc_mapped = TrainerMapping.objects.filter(
                trainer=associate, center=batch.center, course=batch.course, status='active',
            ).exists()
            if not assoc_mapped and not request.user.is_superuser:
                err = 'সহকারী প্রশিক্ষক এই ব্যাচের কেন্দ্র ও কোর্সের জন্য নির্ধারিত নন।'
                return Response({'error': err}, status=400)

        existing_plans = BatchWeekPlan.objects.filter(batch=batch)
        if not existing_plans.exists():
            self._create_default_week_plans(batch, lead, associate)

        updated = BatchWeekPlan.objects.filter(batch=batch).update(
            lead_trainer=lead,
            associate_trainer=associate,
        )
        return Response({
            'updated_plans': updated,
            'lead_trainer_name': lead.user.full_name_bn,
            'lead_trainer_id': lead.id,
            'associate_trainer_name': associate.user.full_name_bn if associate else None,
            'associate_trainer_id': associate.id if associate else None,
        })

    def _create_default_week_plans(self, batch, lead, associate):
        from datetime import time as dt_time
        default_plans = [
            (6, 1, 'theory', '09:00', '11:00', 'তত্ত্ব', 'Theory'),
            (6, 2, 'practical', '11:30', '13:00', 'ব্যবহারিক', 'Practical'),
            (0, 3, 'theory', '09:00', '11:00', 'তত্ত্ব', 'Theory'),
            (0, 4, 'practical', '11:30', '13:00', 'ব্যবহারিক', 'Practical'),
            (1, 5, 'practical', '09:00', '11:00', 'ব্যবহারিক', 'Practical'),
            (2, 6, 'theory', '09:00', '11:00', 'তত্ত্ব', 'Theory'),
            (3, 7, 'practical', '09:00', '11:00', 'ব্যবহারিক', 'Practical'),
            (4, 8, 'theory', '09:00', '11:00', 'তত্ত্ব', 'Theory'),
        ]
        for dow, sno, ct, st, et, t_bn, t_en in default_plans:
            start_time = dt_time(*map(int, st.split(':')))
            end_time = dt_time(*map(int, et.split(':')))
            dur = round((end_time.hour * 60 + end_time.minute - start_time.hour * 60 - start_time.minute) / 60, 1)
            BatchWeekPlan.objects.create(
                batch=batch, term_no=1, term_day=dow, session_no=sno,
                class_type=ct, start_date=batch.start_date, end_date=batch.end_date,
                day_of_week=dow,
                start_time=start_time, end_time=end_time,
                duration_hours=dur,
                training_room_bn='কক্ষ ১০১', training_room_en='Room 101',
                lead_trainer=lead, associate_trainer=associate,
                topic_bn=t_bn, topic_en=t_en,
            )

    @action(detail=True, methods=['post'])
    def add_week_plan(self, request, pk=None):
        batch = self.get_object()
        data = request.data.copy() if hasattr(request.data, 'copy') else dict(request.data)
        if not data.get('start_date') and batch.start_date:
            data['start_date'] = batch.start_date.isoformat()
        if not data.get('end_date') and batch.end_date:
            data['end_date'] = batch.end_date.isoformat()
        serializer = BatchWeekPlanWriteSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer.save(batch=batch)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['post'])
    def bulk_add_week_plans(self, request, pk=None):
        batch = self.get_object()
        data = request.data.get('plans', [])
        if not isinstance(data, list):
            return Response({'error': 'plans একটি তালিকা হতে হবে।'}, status=400)

        created = []
        errors = []
        with transaction.atomic():
            for idx, plan_data in enumerate(data):
                pdata = plan_data.copy() if hasattr(plan_data, 'copy') else dict(plan_data)
                if not pdata.get('start_date') and batch.start_date:
                    pdata['start_date'] = batch.start_date.isoformat()
                if not pdata.get('end_date') and batch.end_date:
                    pdata['end_date'] = batch.end_date.isoformat()
                serializer = BatchWeekPlanWriteSerializer(data=pdata)
                if serializer.is_valid():
                    serializer.save(batch=batch)
                    created.append(serializer.data)
                else:
                    errors.append({'index': idx, 'errors': serializer.errors})

        return Response({
            'created_count': len(created),
            'error_count': len(errors),
            'created': created,
            'errors': errors,
        })

    @action(detail=True, methods=['get'])
    def validate_hours(self, request, pk=None):
        is_valid, planned, course_hours = validate_batch_hours_match_course(pk)
        return Response({
            'is_valid': is_valid,
            'planned_hours': float(planned),
            'course_hours': course_hours,
            'difference': float(abs(course_hours - planned)),
        })

    @action(detail=True, methods=['get'])
    def enrollments(self, request, pk=None):
        batch = self.get_object()
        enrollments = BatchEnrollment.objects.filter(batch=batch).select_related(
            'trainee__user', 'trainee'
        )
        page = self.paginate_queryset(enrollments)
        if page is not None:
            serializer = BatchEnrollmentSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = BatchEnrollmentSerializer(enrollments, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def enroll_trainees(self, request, pk=None):
        batch = self.get_object()
        serializer = BatchEnrollmentBulkSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        trainee_ids = serializer.validated_data['trainee_ids']
        created = []
        skipped = []

        for tid in trainee_ids:
            _, was_created = BatchEnrollment.objects.get_or_create(
                trainee_id=tid,
                batch=batch,
                defaults={'status': BatchEnrollment.EnrollmentStatus.ACTIVE},
            )
            if was_created:
                created.append(tid)
            else:
                skipped.append(tid)

        batch.filled_seats = BatchEnrollment.objects.filter(
            batch=batch, status=BatchEnrollment.EnrollmentStatus.ACTIVE
        ).count()
        batch.save(update_fields=['filled_seats'])

        return Response({
            'enrolled': len(created),
            'already_enrolled': len(skipped),
            'filled_seats': batch.filled_seats,
            'total_seats': batch.total_seats,
        })

    @action(detail=True, methods=['post'])
    def add_trainee(self, request, pk=None):
        batch = self.get_object()
        reg_no = request.data.get('registration_no', '').strip()
        if not reg_no:
            return Response({'error': 'রেজিস্ট্রেশন নম্বর আবশ্যক।'}, status=400)
        try:
            trainee = Trainee.objects.get(registration_no=reg_no, center=batch.center)
        except Trainee.DoesNotExist:
            return Response({'error': f'"{reg_no}" নম্বরের কোনো প্রশিক্ষণার্থী এই কেন্দ্রে পাওয়া যায়নি।'}, status=404)

        if BatchEnrollment.objects.filter(trainee=trainee, batch=batch).exists():
            return Response({'error': 'প্রশিক্ষণার্থী ইতিমধ্যে এই ব্যাচে নথিভুক্ত।'}, status=400)

        if batch.filled_seats >= batch.total_seats:
            return Response({'error': 'ব্যাচে আসন পূর্ণ।'}, status=400)

        enrollment = BatchEnrollment.objects.create(
            trainee=trainee,
            batch=batch,
            status=BatchEnrollment.EnrollmentStatus.ACTIVE,
        )
        batch.filled_seats = BatchEnrollment.objects.filter(
            batch=batch, status=BatchEnrollment.EnrollmentStatus.ACTIVE
        ).count()
        batch.save(update_fields=['filled_seats'])

        return Response(BatchEnrollmentSerializer(enrollment).data, status=201)

    @action(detail=True, methods=['get'])
    def available_trainees(self, request, pk=None):
        batch = self.get_object()
        enrolled_ids = BatchEnrollment.objects.filter(
            batch=batch, status=BatchEnrollment.EnrollmentStatus.ACTIVE
        ).values_list('trainee_id', flat=True)
        trainees = Trainee.objects.filter(center=batch.center).exclude(
            id__in=enrolled_ids
        ).select_related('user')
        page = self.paginate_queryset(trainees)
        if page is not None:
            serializer = TraineeListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = TraineeListSerializer(trainees, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def generate_calendar(self, request, pk=None):
        self.assert_admin_or_ho(request)
        batch = self.get_object()
        from .models import BatchCalendarDay, Holiday as BatchHoliday
        from apps.attendance.models import Attendance, AttendanceSummary

        week_plans = BatchWeekPlan.objects.filter(batch=batch)
        if not week_plans.exists():
            return Response({'error': 'প্রথমে সাপ্তাহিক পরিকল্পনা তৈরি করুন।'}, status=400)

        days_map = {}
        for wp in week_plans:
            dow = wp.day_of_week
            if dow not in days_map:
                days_map[dow] = set()
            days_map[dow].add(wp.session_no)

        holidays = set(
            BatchHoliday.objects.filter(
                date__gte=batch.start_date,
                date__lte=batch.end_date,
            ).values_list('date', flat=True)
        )

        current = batch.start_date
        created_days = 0
        while current <= batch.end_date:
            if current in holidays or current.weekday() in (4, 5):
                current += timedelta(days=1)
                continue
            dow_map = {0: 6, 1: 0, 2: 1, 3: 2, 4: 3, 5: 4, 6: 5}
            py_dow = dow_map[current.weekday()]
            if py_dow in days_map:
                _, was_created = BatchCalendarDay.objects.get_or_create(
                    batch=batch, date=current,
                    defaults={'total_sessions': len(days_map[py_dow]), 'is_generated': True},
                )
                if was_created:
                    created_days += 1
            current += timedelta(days=1)

        active_enrollments = BatchEnrollment.objects.filter(
            batch=batch, status=BatchEnrollment.EnrollmentStatus.ACTIVE,
        ).select_related('trainee')

        cal_days = BatchCalendarDay.objects.filter(batch=batch, is_generated=True, is_holiday=False)
        attendance_created = 0
        for cal_day in cal_days:
            for session_no in range(1, cal_day.total_sessions + 1):
                for enrollment in active_enrollments:
                    _, was_created = Attendance.objects.get_or_create(
                        trainee=enrollment.trainee,
                        batch=batch,
                        session_date=cal_day.date,
                        session_no=session_no,
                        defaults={'status': Attendance.Status.PRESENT},
                    )
                    if was_created:
                        attendance_created += 1

        for enrollment in active_enrollments:
            summary, _ = AttendanceSummary.objects.get_or_create(
                trainee=enrollment.trainee,
                batch=batch,
            )
            summary.refresh()

        return Response({
            'calendar_days_created': created_days,
            'attendance_records_created': attendance_created,
            'enrolled_trainees': active_enrollments.count(),
            'total_calendar_days': cal_days.count(),
        })

    def assert_admin_or_ho(self, request):
        if request.user.user_type not in ('center_admin', 'head_office') and not request.user.is_superuser:
            from rest_framework.exceptions import PermissionDenied
            raise PermissionDenied('শুধুমাত্র কেন্দ্র প্রশাসক বা হেড অফিস এই কাজটি করতে পারেন।')

    @action(detail=True, methods=['post'])
    def start(self, request, pk=None):
        self.assert_admin_or_ho(request)
        batch = self.get_object()
        if batch.status != Batch.BatchStatus.SCHEDULED:
            return Response({'error': 'শুধুমাত্র নির্ধারিত ব্যাচ শুরু করা যাবে।'}, status=400)

        is_valid, planned, course_hours = validate_batch_hours_match_course(pk)
        if not is_valid:
            return Response({
                'error': (
                    f'সাপ্তাহিক পরিকল্পনার মোট ঘন্টা ({float(planned):.1f}) '
                    f'কোর্সের মোট ঘন্টার ({course_hours}) সাথে মিলছে না।'
                ),
                'planned_hours': float(planned),
                'course_hours': course_hours,
            }, status=400)

        batch.status = Batch.BatchStatus.RUNNING
        batch.save(update_fields=['status'])
        return Response(BatchDetailSerializer(batch).data)

    @action(detail=True, methods=['post'])
    def complete(self, request, pk=None):
        self.assert_admin_or_ho(request)
        batch = self.get_object()
        if batch.status not in (Batch.BatchStatus.SCHEDULED, Batch.BatchStatus.RUNNING):
            return Response({'error': 'ব্যাচ ইতিমধ্যে সমাপ্ত বা বাতিল।'}, status=400)

        BatchEnrollment.objects.filter(batch=batch, status=BatchEnrollment.EnrollmentStatus.ACTIVE).update(
            status=BatchEnrollment.EnrollmentStatus.COMPLETED,
        )

        batch.status = Batch.BatchStatus.COMPLETED
        batch.save(update_fields=['status'])
        return Response(BatchDetailSerializer(batch).data)

    @action(detail=True, methods=['post'])
    def cancel(self, request, pk=None):
        self.assert_admin_or_ho(request)
        batch = self.get_object()
        if batch.status == Batch.BatchStatus.COMPLETED:
            return Response({'error': 'সমাপ্ত ব্যাচ বাতিল করা যাবে না।'}, status=400)

        batch.status = Batch.BatchStatus.CANCELLED
        batch.save(update_fields=['status'])
        return Response(BatchDetailSerializer(batch).data)

    @action(detail=False, methods=['get'], url_path='export-list')
    def export_list(self, request):
        qs = self.filter_queryset(self.get_queryset())
        ids = request.query_params.get('ids', '')
        if ids:
            id_list = [int(x) for x in ids.split(',') if x.isdigit()]
            if id_list:
                qs = qs.filter(id__in=id_list)
        fmt = request.query_params.get('file_format', 'xlsx')

        headers = [
            'ব্যাচ নং', 'নাম (বাংলা)', 'নাম (ইংরেজি)', 'কোর্স', 'কেন্দ্র',
            'শুরুর তারিখ', 'সমাপ্তির তারিখ', 'মোট আসন', 'পূরণকৃত', 'অবস্থা',
        ]
        rows = []
        for b in qs:
            rows.append([
                b.batch_no, b.batch_name_bn or '', b.batch_name_en or '',
                b.course.name_bn if b.course else '',
                b.center.name_bn if b.center else '',
                str(b.start_date) if b.start_date else '',
                str(b.end_date) if b.end_date else '',
                b.total_seats or 0, b.filled_seats or 0,
                b.get_status_display() if b.status else '',
            ])

        if fmt == 'xlsx':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Batches'
            ws.append(headers)
            for row in rows:
                ws.append(row)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="batches_{date.today().isoformat()}.xlsx"'
            return response
        else:
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)
            response = HttpResponse(
                output.getvalue().encode('utf-8-sig'),
                content_type='text/csv',
            )
            response['Content-Disposition'] = f'attachment; filename="batches_{date.today().isoformat()}.csv"'
            return response

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def import_list(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'ফাইল নির্বাচন করুন'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            rows_iter = iter(ws.iter_rows(values_only=True))
            header_row = [str(c).strip().lower() if c is not None else '' for c in next(rows_iter)]
        except Exception:
            file.seek(0)
            try:
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                header_row = [h.strip().lower() for h in reader.fieldnames]
                rows_iter = reader
            except Exception:
                return Response({'error': 'ভুল ফাইল ফরম্যাট। Excel (.xlsx) বা CSV ফাইল আপলোড করুন।'}, status=400)

        bn_required = {'ব্যাচ নং'}
        header_set = set(header_row)
        if not bn_required.issubset(header_set):
            return Response({
                'error': 'প্রয়োজনীয় কলাম নেই। হেডারে "ব্যাচ নং" থাকা আবশ্যক।',
                'detected_headers': header_row,
            }, status=400)

        field_map = {
            'ব্যাচ নং': 'batch_no', 'batch_no': 'batch_no',
            'নাম (বাংলা)': 'batch_name_bn', 'batch_name_bn': 'batch_name_bn',
            'নাম (ইংরেজি)': 'batch_name_en', 'batch_name_en': 'batch_name_en',
            'অবস্থা': 'status', 'status': 'status',
            'মোট আসন': 'total_seats', 'total_seats': 'total_seats',
        }

        results = {'created': 0, 'updated': 0, 'errors': []}
        for row_idx, row in enumerate(rows_iter, start=2):
            try:
                if isinstance(row, dict):
                    raw = row
                else:
                    raw = dict(zip(header_row, [str(c).strip() if c is not None else '' for c in row]))

                data = {}
                for k, v in raw.items():
                    mapped = field_map.get(k.strip().lower(), k.strip().lower())
                    data[mapped] = v.strip() if v else ''

                batch_no = data.get('batch_no', '').strip()
                if not batch_no:
                    results['errors'].append(f'সারি {row_idx}: ব্যাচ নং আবশ্যক')
                    continue

                existing = Batch.objects.filter(batch_no=batch_no).first()
                if existing:
                    if data.get('batch_name_bn') is not None:
                        existing.batch_name_bn = data.get('batch_name_bn') or ''
                    if data.get('batch_name_en') is not None:
                        existing.batch_name_en = data.get('batch_name_en') or ''
                    if data.get('total_seats') is not None:
                        try:
                            existing.total_seats = int(data['total_seats']) if data['total_seats'] else None
                        except ValueError:
                            pass
                    if data.get('status') is not None:
                        existing.status = data['status']
                    existing.save()
                    results['updated'] += 1
                else:
                    results['errors'].append(f'সারি {row_idx}: "{batch_no}" ব্যাচ নং পাওয়া যায়নি')
            except Exception as e:
                results['errors'].append(f'সারি {row_idx}: {str(e)}')

        return Response(results)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Template'
        headers = ['ব্যাচ নং', 'নাম (বাংলা)', 'নাম (ইংরেজি)', 'মোট আসন', 'অবস্থা']
        ws.append(headers)
        sample = ['', 'উদাহরণ ব্যাচ', 'Example Batch', '৩০', 'scheduled']
        ws.append(sample)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = openpyxl.styles.Font(bold=True)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="batch_import_template.xlsx"'
        wb.save(response)
        return response

    @action(detail=False, methods=['post'])
    def generate(self, request):
        circular_id = request.data.get('circular_id')
        if not circular_id:
            return Response({'error': 'circular_id is required'}, status=status.HTTP_400_BAD_REQUEST)

        from apps.circulars.models import CircularCenterAllocation, Circular
        from apps.applications.models import Application
        from django.db import transaction

        user = request.user
        if user.user_type == 'center_admin' and user.center:
            center = user.center
        else:
            center_id = request.data.get('center_id')
            if not center_id:
                return Response({'error': 'center_id required for non-center-admin'}, status=400)
            center = Center.objects.get(id=center_id)

        try:
            allocation = CircularCenterAllocation.objects.get(
                circular_id=circular_id, center=center,
            )
        except CircularCenterAllocation.DoesNotExist:
            return Response(
                {'error': 'No seat allocation found for this circular and center'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        circular = Circular.objects.get(id=circular_id)
        batch_size = 25
        num_batches = allocation.allocated_seats // batch_size
        if num_batches == 0:
            return Response({'error': 'Allocated seats too few for even one batch'}, status=400)

        selected_applicants = list(
            Application.objects.filter(
                circular_id=circular_id,
                chosen_center=center,
                status=Application.ApplicationStatus.SELECTED,
            ).select_related('user__trainee_profile')
        )
        if not selected_applicants:
            return Response({'error': 'কোনো নির্বাচিত আবেদনকারী নেই'}, status=400)

        existing = Batch.objects.filter(circular_id=circular_id, center=center).count()
        created = []
        with transaction.atomic():
            for i in range(num_batches):
                idx = existing + i + 1
                shift = Batch.ShiftChoice.SHIFT_1 if idx <= num_batches / 2 else Batch.ShiftChoice.SHIFT_2
                batch = Batch.objects.create(
                    circular=circular,
                    center=center,
                    course=circular.course,
                    batch_name_bn=f'{circular.title_bn} - ব্যাচ {idx}',
                    batch_name_en=f'{circular.title_en} - Batch {idx}',
                    shift=shift,
                    batch_index=idx,
                    start_date=circular.training_start_date,
                    end_date=circular.training_end_date,
                    total_seats=batch_size,
                    created_by=request.user,
                )
                # Enroll selected applicants into this batch (up to batch_size)
                chunk = selected_applicants[i * batch_size:(i + 1) * batch_size]
                for app in chunk:
                    from apps.trainees.models import Trainee
                    trainee, created = Trainee.objects.get_or_create(
                        user=app.user,
                        defaults={
                            'application': app,
                            'center': app.chosen_center,
                        },
                    )
                    if not created:
                        changed = False
                        if trainee.center_id != app.chosen_center_id:
                            trainee.center = app.chosen_center
                            changed = True
                        if trainee.application_id != app.id:
                            trainee.application = app
                            changed = True
                        if changed:
                            trainee.save(update_fields=['center', 'application'])
                    BatchEnrollment.objects.create(
                            trainee=trainee,
                            batch=batch,
                            status=BatchEnrollment.EnrollmentStatus.ACTIVE,
                        )
                    if trainee.batch_id != batch.id:
                        trainee.batch = batch
                        trainee.save(update_fields=['batch'])
                    app.status = Application.ApplicationStatus.ENROLLED
                    app.save(update_fields=['status'])

                batch.filled_seats = len(chunk)
                batch.save(update_fields=['filled_seats'])
                created.append(BatchDetailSerializer(batch).data)

        return Response({'batches': created, 'count': len(created)}, status=status.HTTP_201_CREATED)


class BatchWeekPlanViewSet(viewsets.ModelViewSet):
    queryset = BatchWeekPlan.objects.select_related(
        'batch', 'lead_trainer__user', 'associate_trainer__user'
    ).all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields = ('batch', 'term_no', 'class_type', 'day_of_week', 'lead_trainer')
    ordering = ('batch', 'term_no', 'term_day', 'session_no')

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return BatchWeekPlanWriteSerializer
        return BatchWeekPlanListSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.user_type == 'center_admin' and user.center:
            qs = qs.filter(batch__center=user.center)
        elif user.user_type == 'center_admin' and not user.center:
            return BatchWeekPlan.objects.none()
        return qs


class BatchEnrollmentViewSet(viewsets.ModelViewSet):
    queryset = BatchEnrollment.objects.select_related(
        'trainee__user', 'batch'
    ).all()
    permission_classes = [IsAuthenticated]
    serializer_class = BatchEnrollmentSerializer
    filter_backends = [DjangoFilterBackend, filters.SearchFilter]
    filterset_fields = ('batch', 'status')
    search_fields = ('trainee__registration_no', 'trainee__user__full_name_bn', 'trainee__user__phone')
    ordering = ('-enrollment_date',)

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.user_type == 'center_admin' and user.center:
            qs = qs.filter(batch__center=user.center)
        elif user.user_type == 'center_admin' and not user.center:
            return BatchEnrollment.objects.none()
        return qs

    @action(detail=True, methods=['post'])
    def drop(self, request, pk=None):
        enrollment = self.get_object()
        if enrollment.status != BatchEnrollment.EnrollmentStatus.ACTIVE:
            return Response({'error': 'শুধুমাত্র সক্রিয় নথিভুক্তি বাতিল করা যাবে।'}, status=400)

        from django.utils import timezone
        enrollment.status = BatchEnrollment.EnrollmentStatus.DROPPED
        enrollment.dropped_date = request.data.get('dropped_date') or timezone.now().date()
        enrollment.drop_reason = request.data.get('drop_reason', '')
        enrollment.save(update_fields=['status', 'dropped_date', 'drop_reason'])

        batch = enrollment.batch
        batch.filled_seats = BatchEnrollment.objects.filter(
            batch=batch, status=BatchEnrollment.EnrollmentStatus.ACTIVE
        ).count()
        batch.save(update_fields=['filled_seats'])

        return Response(BatchEnrollmentSerializer(enrollment).data)

    @action(detail=True, methods=['post'])
    def transfer(self, request, pk=None):
        enrollment = self.get_object()
        if enrollment.status != BatchEnrollment.EnrollmentStatus.ACTIVE:
            return Response({'error': 'শুধুমাত্র সক্রিয় নথিভুক্তি স্থানান্তর করা যাবে।'}, status=400)

        target_batch_id = request.data.get('target_batch_id')
        if not target_batch_id:
            return Response({'error': 'target_batch_id আবশ্যক।'}, status=400)

        try:
            target_batch = Batch.objects.get(id=target_batch_id)
        except Batch.DoesNotExist:
            return Response({'error': 'টার্গেট ব্যাচ পাওয়া যায়নি।'}, status=404)

        if target_batch.center_id != enrollment.batch.center_id:
            return Response({'error': 'একই কেন্দ্রের ব্যাচে স্থানান্তর করা যাবে।'}, status=400)

        if target_batch.filled_seats >= target_batch.total_seats:
            return Response({'error': 'টার্গেট ব্যাচে আসন পূর্ণ।'}, status=400)

        if BatchEnrollment.objects.filter(trainee=enrollment.trainee, batch=target_batch).exists():
            return Response({'error': 'প্রশিক্ষণার্থী ইতিমধ্যে টার্গেট ব্যাচে নথিভুক্ত।'}, status=400)

        from django.utils import timezone
        old_batch = enrollment.batch

        enrollment.status = BatchEnrollment.EnrollmentStatus.DROPPED
        enrollment.dropped_date = timezone.now().date()
        enrollment.drop_reason = 'স্থানান্তর'
        enrollment.save(update_fields=['status', 'dropped_date', 'drop_reason'])

        new_enrollment = BatchEnrollment.objects.create(
            trainee=enrollment.trainee,
            batch=target_batch,
            status=BatchEnrollment.EnrollmentStatus.ACTIVE,
        )

        old_batch.filled_seats = BatchEnrollment.objects.filter(
            batch=old_batch, status=BatchEnrollment.EnrollmentStatus.ACTIVE
        ).count()
        old_batch.save(update_fields=['filled_seats'])

        target_batch.filled_seats = BatchEnrollment.objects.filter(
            batch=target_batch, status=BatchEnrollment.EnrollmentStatus.ACTIVE
        ).count()
        target_batch.save(update_fields=['filled_seats'])

        return Response(BatchEnrollmentSerializer(new_enrollment).data, status=201)


class ShiftViewSet(viewsets.ModelViewSet):
    from .models import Shift
    queryset = Shift.objects.all()
    permission_classes = [IsAuthenticated]
    serializer_class = ShiftSerializer
    filterset_fields = ('center', 'is_active')

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.user_type == 'center_admin' and user.center:
            qs = qs.filter(Q(center=user.center) | Q(center__isnull=True))
        return qs

    def perform_create(self, serializer):
        user = self.request.user
        if user.user_type == 'center_admin' and user.center:
            serializer.save(center=user.center)
        else:
            serializer.save()

    def _assert_own_center(self, instance):
        user = self.request.user
        if user.user_type == 'center_admin':
            if instance.center is None or instance.center != user.center:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied('শুধুমাত্র নিজ কেন্দ্রের শিফট সম্পাদনা/মুছতে পারবেন')

    def perform_update(self, serializer):
        self._assert_own_center(serializer.instance)
        serializer.save()

    def perform_destroy(self, instance):
        self._assert_own_center(instance)
        instance.delete()


class HolidayViewSet(viewsets.ModelViewSet):
    from .models import Holiday
    queryset = Holiday.objects.all()
    permission_classes = [IsAuthenticated]
    serializer_class = HolidaySerializer
    filterset_fields = ('center', 'is_government_holiday')
    ordering = ('-date',)

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.user_type == 'center_admin' and user.center:
            qs = qs.filter(Q(center=user.center) | Q(center__isnull=True))
        return qs

    def perform_create(self, serializer):
        user = self.request.user
        if user.user_type == 'center_admin' and user.center:
            serializer.save(center=user.center)

    def _assert_own_center(self, instance):
        user = self.request.user
        if user.user_type == 'center_admin':
            if instance.center is None or instance.center != user.center:
                from rest_framework.exceptions import PermissionDenied
                raise PermissionDenied('শুধুমাত্র নিজ কেন্দ্রের ছুটির দিন সম্পাদনা/মুছতে পারবেন')

    def perform_update(self, serializer):
        self._assert_own_center(serializer.instance)
        serializer.save()

    def perform_destroy(self, instance):
        self._assert_own_center(instance)
        instance.delete()

    @action(detail=False, methods=['post'])
    def seed(self, request):
        from .models import Holiday
        year = request.data.get('year')
        from datetime import date
        today = date.today()
        if not year:
            year = today.year
        year = int(year)

        fixed_holidays = [
            (f'{year}-02-21', 'শহীদ দিবস ও আন্তর্জাতিক মাতৃভাষা দিবস', 'Shaheed Day & International Mother Language Day', True),
            (f'{year}-03-17', 'বঙ্গবন্ধুর জন্মদিন', 'Bangabandhu Sheikh Mujibur Rahman\'s Birthday', True),
            (f'{year}-03-26', 'স্বাধীনতা ও জাতীয় দিবস', 'Independence & National Day', True),
            (f'{year}-04-14', 'পহেলা বৈশাখ', 'Bengali New Year (Pohela Boishakh)', True),
            (f'{year}-05-01', 'মে দিবস', 'May Day', True),
            (f'{year}-08-15', 'জাতীয় শোক দিবস', 'National Mourning Day', True),
            (f'{year}-12-16', 'বিজয় দিবস', 'Victory Day', True),
            (f'{year}-12-25', 'বড়দিন', 'Christmas Day', False),
        ]

        created = []
        skipped = 0
        for date_str, desc_bn, desc_en, is_gov in fixed_holidays:
            dt = date.fromisoformat(date_str)
            if Holiday.objects.filter(date=dt).exists():
                skipped += 1
                continue
            Holiday.objects.create(
                date=dt,
                description_bn=desc_bn,
                description_en=desc_en,
                is_government_holiday=is_gov,
            )
            created.append(date_str)

        return Response({
            'created': created,
            'skipped': skipped,
            'year': year,
            'message': f'{year} সালে {len(created)}টি ছুটির দিন যোগ করা হয়েছে, {skipped}টি ইতিমধ্যে বিদ্যমান',
        })

    @action(detail=False, methods=['get'])
    def is_working_day(self, request):
        from .models import Holiday
        from datetime import date
        d = request.query_params.get('date')
        if d:
            from datetime import datetime
            d = datetime.strptime(d, '%Y-%m-%d').date()
        else:
            d = date.today()
        # Weekend check
        if d.weekday() in (4, 5):  # Friday=4, Saturday=5 in Python weekday()
            return Response({'date': d, 'is_working': False, 'reason': 'Weekend'})
        # Holiday check
        if Holiday.objects.filter(date=d).exists():
            return Response({'date': d, 'is_working': False, 'reason': 'Holiday'})
        return Response({'date': d, 'is_working': True, 'reason': 'Working day'})

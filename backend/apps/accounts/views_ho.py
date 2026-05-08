import io
import csv
import secrets
import string
from datetime import date

import openpyxl
from django.db import transaction
from django.http import HttpResponse
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets, permissions, status, filters as drf_filters
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response

from apps.centers.models import ActionLog, Center
from .models import User, UserProfile, Role, LoginLog
from .serializers_ho import (
    HOUserListSerializer, HOUserDetailSerializer, HOUserCreateSerializer,
    HOUserUpdateSerializer, HOUserImportSerializer,
    LoginLogSerializer, ActionLogSerializer,
    RoleSerializer, RolePermissionSerializer,
)


class IsHeadOffice(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        return request.user.user_type == 'head_office' or request.user.is_superuser


AVAILABLE_PERMISSIONS = [
    {
        'key': 'centers',
        'label': 'Center Permissions',
        'children': [
            {'key': 'centers.view', 'label': 'View Centers'},
            {'key': 'centers.create', 'label': 'Create Centers'},
            {'key': 'centers.edit', 'label': 'Edit Centers'},
            {'key': 'centers.delete', 'label': 'Delete Centers'},
        ],
    },
    {
        'key': 'courses',
        'label': 'Course Permissions',
        'children': [
            {'key': 'courses.view', 'label': 'View Courses'},
            {'key': 'courses.create', 'label': 'Create Courses'},
            {'key': 'courses.edit', 'label': 'Edit Courses'},
            {'key': 'courses.delete', 'label': 'Delete Courses'},
        ],
    },
    {
        'key': 'trainers',
        'label': 'Trainer Permissions',
        'children': [
            {'key': 'trainers.view', 'label': 'View Trainers'},
            {'key': 'trainers.approve', 'label': 'Approve Trainers'},
            {'key': 'trainers.suspend', 'label': 'Suspend Trainers'},
            {'key': 'trainers.map', 'label': 'Map Trainers to Centers'},
        ],
    },
    {
        'key': 'assessors',
        'label': 'Assessor Permissions',
        'children': [
            {'key': 'assessors.view', 'label': 'View Assessors'},
            {'key': 'assessors.approve', 'label': 'Approve Assessors'},
            {'key': 'assessors.convert', 'label': 'Convert Trainers to Assessors'},
        ],
    },
    {
        'key': 'trainees',
        'label': 'Trainee Permissions',
        'children': [
            {'key': 'trainees.view', 'label': 'View Trainees'},
            {'key': 'trainees.edit', 'label': 'Edit Trainees'},
            {'key': 'trainees.enroll', 'label': 'Enroll Trainees'},
        ],
    },
    {
        'key': 'financial',
        'label': 'Financial Permissions',
        'children': [
            {'key': 'financial.view_budget', 'label': 'View Budget'},
            {'key': 'financial.edit_budget', 'label': 'Edit Budget'},
            {'key': 'financial.create_voucher', 'label': 'Create Vouchers'},
            {'key': 'financial.verify_voucher', 'label': 'Verify Vouchers'},
            {'key': 'financial.approve_voucher', 'label': 'Approve Vouchers'},
        ],
    },
    {
        'key': 'reports',
        'label': 'Report Permissions',
        'children': [
            {'key': 'reports.view', 'label': 'View Reports'},
            {'key': 'reports.export', 'label': 'Export Reports'},
            {'key': 'reports.schedule', 'label': 'Schedule Reports'},
        ],
    },
    {
        'key': 'users',
        'label': 'User Management',
        'children': [
            {'key': 'users.view', 'label': 'View Users'},
            {'key': 'users.create', 'label': 'Create Users'},
            {'key': 'users.edit', 'label': 'Edit Users'},
            {'key': 'users.delete', 'label': 'Delete Users'},
            {'key': 'users.manage_roles', 'label': 'Manage Roles'},
        ],
    },
]


class HOUserViewSet(viewsets.ModelViewSet):
    queryset = User.objects.select_related('center', 'role', 'profile').all()
    permission_classes = [permissions.IsAuthenticated, IsHeadOffice]
    filter_backends = (DjangoFilterBackend, drf_filters.SearchFilter, drf_filters.OrderingFilter)
    filterset_fields = ('user_type', 'center', 'is_active', 'role')
    search_fields = ('email', 'full_name_bn', 'full_name_en', 'phone', 'nid')
    ordering_fields = ('created_at', 'last_login', 'email', 'full_name_bn')
    ordering = ('-created_at',)

    def get_serializer_class(self):
        if self.action == 'list':
            return HOUserListSerializer
        if self.action == 'create':
            return HOUserCreateSerializer
        if self.action in ('update', 'partial_update'):
            return HOUserUpdateSerializer
        if self.action == 'import_users':
            return HOUserImportSerializer
        return HOUserDetailSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        export = self.request.query_params.get('export')
        if export:
            return qs
        return qs

    def perform_create(self, serializer):
        user = serializer.save()
        password = getattr(user, '_generated_password', None)
        send_email = getattr(user, '_send_email', True)

        ActionLog.objects.create(
            user=self.request.user,
            action='created user',
            target_type='User',
            target_id=str(user.id),
            description=f'User created: {user.email} ({user.get_user_type_display()})',
            ip_address=self.request.META.get('REMOTE_ADDR', ''),
        )

        if send_email and password:
            try:
                from apps.notifications.models import Notification
                from apps.notifications.tasks import send_notification
                notif = Notification.objects.create(
                    recipient=user,
                    subject='Welcome to BRTC TMS - Your Account Credentials',
                    message=(
                        f'Dear {user.full_name_bn},\n\n'
                        f'Your account has been created on BRTC Training Management System.\n\n'
                        f'Email: {user.email}\n'
                        f'Password: {password}\n\n'
                        f'Please login at the BRTC TMS portal and change your password.\n\n'
                        f'Regards,\nBRTC Management'
                    ),
                    channel='email',
                )
                send_notification.delay(notif.id)
            except Exception:
                pass

    def perform_update(self, serializer):
        user = serializer.save()
        ActionLog.objects.create(
            user=self.request.user,
            action='updated user',
            target_type='User',
            target_id=str(user.id),
            description=f'User updated: {user.email}',
            ip_address=self.request.META.get('REMOTE_ADDR', ''),
        )

    def perform_destroy(self, instance):
        ActionLog.objects.create(
            user=self.request.user,
            action='deleted user',
            target_type='User',
            target_id=str(instance.id),
            description=f'User deleted: {instance.email}',
            ip_address=self.request.META.get('REMOTE_ADDR', ''),
        )
        instance.delete()

    @action(detail=True, methods=['post'])
    def reset_password(self, request, pk=None):
        user = self.get_object()
        new_password = request.data.get('password', '')
        if not new_password or len(new_password) < 8:
            new_password = ''.join(secrets.choice(string.ascii_letters + string.digits) for _ in range(12))
        user.set_password(new_password)
        user.save(update_fields=['password'])
        ActionLog.objects.create(
            user=request.user,
            action='reset password',
            target_type='User',
            target_id=str(user.id),
            description=f'Password reset for {user.email}',
            ip_address=request.META.get('REMOTE_ADDR', ''),
        )
        return Response({'new_password': new_password, 'detail': 'পাসওয়ার্ড রিসেট করা হয়েছে'})

    @action(detail=True, methods=['post'])
    def toggle_status(self, request, pk=None):
        user = self.get_object()
        user.is_active = not user.is_active
        user.save(update_fields=['is_active'])
        status_text = 'activated' if user.is_active else 'suspended'
        ActionLog.objects.create(
            user=request.user,
            action=f'{status_text} user',
            target_type='User',
            target_id=str(user.id),
            description=f'User {status_text}: {user.email}',
            ip_address=request.META.get('REMOTE_ADDR', ''),
        )
        return Response({
            'is_active': user.is_active,
            'detail': 'ব্যবহারকারী সক্রিয় করা হয়েছে' if user.is_active else 'ব্যবহারকারী নিষ্ক্রিয় করা হয়েছে',
        })

    @action(detail=True, methods=['post'])
    def reset_mfa(self, request, pk=None):
        user = self.get_object()
        user.mfa_secret = None
        user.save(update_fields=['mfa_secret'])
        ActionLog.objects.create(
            user=request.user,
            action='reset mfa',
            target_type='User',
            target_id=str(user.id),
            description=f'MFA reset for {user.email}',
            ip_address=request.META.get('REMOTE_ADDR', ''),
        )
        return Response({'detail': 'MFA রিসেট করা হয়েছে'})

    @action(detail=True, methods=['get'])
    def login_history(self, request, pk=None):
        user = self.get_object()
        logs = LoginLog.objects.filter(user=user).order_by('-login_time')[:50]
        serializer = LoginLogSerializer(logs, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['get'])
    def audit_log(self, request, pk=None):
        logs = ActionLog.objects.filter(
            target_type='User', target_id=str(pk),
        ).select_related('user').order_by('-created_at')[:50]
        serializer = ActionLogSerializer(logs, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def export(self, request):
        qs = self.filter_queryset(self.get_queryset())
        fmt = request.query_params.get('format', 'csv')

        headers = [
            'Email', 'Name (BN)', 'Name (EN)', 'User Type', 'Center',
            'Phone', 'NID', 'Role', 'Status', 'Last Login', 'Created',
        ]
        rows = []
        for u in qs:
            rows.append([
                u.email, u.full_name_bn, u.full_name_en,
                u.get_user_type_display(),
                u.center.name_bn if u.center else '',
                u.phone, u.nid,
                u.role.name if u.role else '',
                'Active' if u.is_active else 'Suspended',
                u.last_login.isoformat() if u.last_login else '',
                u.created_at.isoformat() if u.created_at else '',
            ])

        if fmt == 'excel':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Users'
            ws.append(headers)
            for row in rows:
                ws.append(row)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="users_{date.today().isoformat()}.xlsx"'
            return response
        else:
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)
            response = HttpResponse(
                output.getvalue().encode('utf-8-sig'),
                content_type='text/csv',
            )
            response['Content-Disposition'] = f'attachment; filename="users_{date.today().isoformat()}.csv"'
            return response

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def import_users(self, request):
        serializer = HOUserImportSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        file = serializer.validated_data['file']
        send_welcome = serializer.validated_data.get('send_welcome', True)
        update_existing = serializer.validated_data.get('update_existing', False)

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            rows_iter = iter(ws.iter_rows(values_only=True))
            header_row = [str(c).strip().lower() if c else '' for c in next(rows_iter)]
        except Exception:
            file.seek(0)
            try:
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                header_row = [h.strip().lower() for h in reader.fieldnames]
                rows_iter = reader
            except Exception:
                return Response({'error': 'Invalid file format. Upload Excel (.xlsx) or CSV.'}, status=400)

        required = {'email', 'full_name_bn', 'full_name_en', 'phone', 'nid'}
        missing = required - set(header_row)
        if missing:
            return Response({
                'error': f'Missing columns: {", ".join(sorted(missing))}',
            }, status=400)

        results = {'created': 0, 'updated': 0, 'errors': []}
        centers = {c.code: c.id for c in Center.objects.all()}
        roles = {r.name: r.id for r in Role.objects.all()}

        for row_idx, row in enumerate(rows_iter, start=2):
            try:
                if isinstance(row, dict):
                    data = row
                else:
                    data = dict(zip(header_row, [str(c) if c else '' for c in row]))

                email = data.get('email', '').strip()
                if not email:
                    continue

                user_data = {
                    'email': email,
                    'full_name_bn': data.get('full_name_bn', '').strip(),
                    'full_name_en': data.get('full_name_en', '').strip(),
                    'phone': str(data.get('phone', '')).strip(),
                    'nid': str(data.get('nid', '')).strip(),
                    'birth_certificate_no': str(data.get('birth_certificate_no', '')).strip(),
                    'user_type': data.get('user_type', 'trainee').strip(),
                }

                center_code = data.get('center', '').strip()
                if center_code and center_code in centers:
                    user_data['center_id'] = centers[center_code]

                role_name = data.get('role', '').strip()
                if role_name and role_name in roles:
                    user_data['role_id'] = roles[role_name]

                existing = User.objects.filter(email=email).first()
                if existing:
                    if update_existing:
                        for k, v in user_data.items():
                            setattr(existing, k, v)
                        existing.save()
                        results['updated'] += 1
                    else:
                        results['errors'].append(f'Row {row_idx}: {email} already exists')
                else:
                    password = data.get('password', '').strip() or ''.join(
                        [secrets.choice(string.ascii_letters + string.digits) for _ in range(12)],
                    )
                    profile_fields = {}
                    for pf in ('gender', 'date_of_birth', 'present_address', 'permanent_address'):
                        if pf in data and data[pf]:
                            profile_fields[pf] = data[pf]

                    user = User.objects.create_user(
                        email=user_data.pop('email'),
                        password=password,
                        **user_data,
                    )
                    if profile_fields:
                        UserProfile.objects.update_or_create(user=user, defaults=profile_fields)
                    results['created'] += 1

                    if send_welcome:
                        try:
                            from apps.notifications.models import Notification
                            from apps.notifications.tasks import send_notification
                            notif = Notification.objects.create(
                                recipient=user,
                                subject='Welcome to BRTC TMS - Your Account Credentials',
                                message=(
                                    f'Dear {user.full_name_bn},\n\n'
                                    f'Your account has been created.\n'
                                    f'Email: {user.email}\nPassword: {password}\n\n'
                                    f'Regards,\nBRTC Management'
                                ),
                                channel='email',
                            )
                            send_notification.delay(notif.id)
                        except Exception:
                            pass
            except Exception as e:
                results['errors'].append(f'Row {row_idx}: {str(e)}')

        ActionLog.objects.create(
            user=request.user,
            action='imported users',
            target_type='User',
            description=f'Imported users: {results["created"]} created, {results["updated"]} updated',
            ip_address=request.META.get('REMOTE_ADDR', ''),
        )
        return Response(results)


class HORoleViewSet(viewsets.ModelViewSet):
    queryset = Role.objects.all()
    permission_classes = [permissions.IsAuthenticated, IsHeadOffice]
    serializer_class = RoleSerializer
    filter_backends = (drf_filters.SearchFilter,)
    search_fields = ('name',)
    ordering = ('name',)

    def perform_create(self, serializer):
        role = serializer.save()
        ActionLog.objects.create(
            user=self.request.user,
            action='created role',
            target_type='Role',
            target_id=str(role.id),
            description=f'Role created: {role.name}',
            ip_address=self.request.META.get('REMOTE_ADDR', ''),
        )

    def perform_destroy(self, instance):
        if instance.is_system:
            return Response({'error': 'System roles cannot be deleted'}, status=400)
        instance.delete()

    @action(detail=True, methods=['put'])
    def permissions(self, request, pk=None):
        role = self.get_object()
        perms = request.data.get('permissions', [])
        role.permissions = perms
        role.save(update_fields=['permissions'])
        ActionLog.objects.create(
            user=request.user,
            action='updated role permissions',
            target_type='Role',
            target_id=str(role.id),
            description=f'Permissions updated for role: {role.name}',
            ip_address=request.META.get('REMOTE_ADDR', ''),
        )
        return Response(RoleSerializer(role).data)

    @action(detail=False, methods=['get'])
    def available_permissions(self, request):
        serializer = RolePermissionSerializer(AVAILABLE_PERMISSIONS, many=True)
        return Response(serializer.data)

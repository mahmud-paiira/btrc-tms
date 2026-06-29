import csv
import io
from datetime import date

import openpyxl
from django.utils import timezone
from rest_framework import viewsets, filters, status, permissions
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from .models import Trainer, TrainerMapping
from .serializers import (
    TrainerListSerializer,
    TrainerDetailSerializer,
    TrainerWriteSerializer,
    TrainerMappingSerializer,
    TrainerApprovalSerializer,
    TrainerTrackSerializer,
)


class IsCenterAdmin(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.user_type == 'center_admin' or request.user.is_superuser


class TrainerViewSet(viewsets.ModelViewSet):
    queryset = Trainer.objects.select_related('user', 'approved_by').prefetch_related(
        'mappings__center', 'mappings__course',
    ).all()
    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)
    filterset_fields = ('status', 'approval_status', 'expertise_area')

    def get_permissions(self):
        if self.action == 'list':
            return [permissions.IsAuthenticated()]
        return [permissions.IsAuthenticated(), IsCenterAdmin()]
    search_fields = (
        'trainer_no', 'nid', 'birth_certificate_no',
        'user__email', 'user__phone',
        'user__full_name_bn', 'user__full_name_en',
    )
    ordering_fields = ('trainer_no', 'years_of_experience', 'created_at')
    ordering = ('-created_at',)

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        center_id = self.request.query_params.get('mapped_center')
        course_id = self.request.query_params.get('mapped_course')

        if center_id:
            qs = qs.filter(mappings__center_id=center_id, mappings__status='active')
        elif user.user_type == 'center_admin' and user.center:
            qs = qs.filter(user__center=user.center)

        if course_id:
            qs = qs.filter(mappings__course_id=course_id, mappings__status='active')

        if center_id or course_id:
            qs = qs.distinct()
        return qs

    def perform_create(self, serializer):
        user = self.request.user
        center = user.center if user.user_type == 'center_admin' and user.center else None
        seq = 1
        if center:
            prefix = f'C{center.id}-'
            last = Trainer.objects.filter(trainer_no__startswith=prefix).order_by('trainer_no').last()
            if last:
                seq = int(last.trainer_no.split('-')[-1]) + 1
        trainer_no = f'C{center.id}-{seq}' if center else str(seq)
        user_id = self.request.data.get('user')
        trainer = serializer.save(user_id=user_id, trainer_no=trainer_no)
        if center:
            trainer.user.center = center
            trainer.user.save(update_fields=['center'])

    def perform_update(self, serializer):
        trainer = serializer.save()
        user = trainer.user
        data = self.request.data
        allowed = {k: data[k] for k in ('full_name_bn', 'full_name_en', 'phone') if k in data}
        if allowed:
            for k, v in allowed.items():
                setattr(user, k, v)
            user.save(update_fields=list(allowed.keys()))

    def get_serializer_class(self):
        if self.action == 'list':
            return TrainerListSerializer
        if self.action in ('create', 'update', 'partial_update'):
            return TrainerWriteSerializer
        return TrainerDetailSerializer

    @action(detail=False, methods=['post'])
    def track(self, request):
        serializer = TrainerTrackSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        query = serializer.validated_data['query']
        search_by = serializer.validated_data['search_by']

        lookup = {
            'trainer_no': 'trainer_no__iexact',
            'nid': 'nid__iexact',
            'mobile': 'user__phone__iexact',
            'birth_certificate_no': 'birth_certificate_no__iexact',
        }
        try:
            trainer = Trainer.objects.select_related('user').get(
                **{lookup[search_by]: query}
            )
            return Response(TrainerDetailSerializer(trainer).data)
        except Trainer.DoesNotExist:
            return Response(
                {'error': 'কোনো প্রশিক্ষক পাওয়া যায়নি'},
                status=status.HTTP_404_NOT_FOUND,
            )

    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        trainer = self.get_object()
        serializer = TrainerApprovalSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        action_val = serializer.validated_data['action']
        if action_val == 'approve':
            trainer.approval_status = Trainer.ApprovalStatus.APPROVED
            trainer.status = Trainer.Status.ACTIVE
        else:
            trainer.approval_status = Trainer.ApprovalStatus.REJECTED
            trainer.status = Trainer.Status.PENDING

        trainer.approved_by = request.user
        trainer.approved_at = timezone.now()
        trainer.save()
        return Response(TrainerDetailSerializer(trainer).data)

    @action(detail=False, methods=['get'], url_path='export-list')
    def export_list(self, request):
        qs = self.filter_queryset(self.get_queryset())
        ids = request.query_params.get('ids', '')
        if ids:
            id_list = [int(x) for x in ids.split(',') if x.isdigit()]
            if id_list:
                qs = qs.filter(id__in=id_list)
        fmt = request.query_params.get('file_format', 'xlsx')

        headers = [
            'প্রশিক্ষক নং', 'নাম (বাংলা)', 'নাম (ইংরেজি)', 'ইমেইল', 'ফোন',
            'এনআইডি', 'শিক্ষাগত যোগ্যতা', 'অভিজ্ঞতা (বছর)', 'দক্ষতার ক্ষেত্র',
            'স্ট্যাটাস', 'অনুমোদন স্ট্যাটাস', 'তৈরির তারিখ',
        ]
        rows = []
        for t in qs:
            rows.append([
                t.trainer_no, t.user.full_name_bn, t.user.full_name_en,
                t.user.email, t.user.phone, t.nid,
                t.education.name_bn if t.education else (t.education_qualification or ''),
                t.years_of_experience or '', t.expertise_area or '',
                t.status, t.approval_status,
                t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '',
            ])

        if fmt == 'xlsx':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Trainers'
            ws.append(headers)
            for row in rows:
                ws.append(row)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="trainers_{date.today().isoformat()}.xlsx"'
            return response
        else:
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)
            response = HttpResponse(
                output.getvalue().encode('utf-8-sig'),
                content_type='text/csv',
            )
            response['Content-Disposition'] = f'attachment; filename="trainers_{date.today().isoformat()}.csv"'
            return response

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def import_list(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'ফাইল নির্বাচন করুন'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            rows_iter = iter(ws.iter_rows(values_only=True))
            header_row = [str(c).strip().lower() if c is not None else '' for c in next(rows_iter)]
        except Exception:
            file.seek(0)
            try:
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                header_row = [h.strip().lower() for h in reader.fieldnames]
                rows_iter = reader
            except Exception:
                return Response({'error': 'ভুল ফাইল ফরম্যাট। Excel (.xlsx) বা CSV ফাইল আপলোড করুন।'}, status=400)

        bn_required = {'প্রশিক্ষক নং'}
        header_set = set(header_row)
        if not bn_required.issubset(header_set):
            return Response({
                'error': f'প্রয়োজনীয় কলাম নেই। হেডারে "প্রশিক্ষক নং" থাকা আবশ্যক।',
                'detected_headers': header_row,
            }, status=400)

        field_map = {
            'প্রশিক্ষক নং': 'trainer_no', 'trainer_no': 'trainer_no',
            'নাম (বাংলা)': 'full_name_bn', 'name_bn': 'full_name_bn',
            'নাম (ইংরেজি)': 'full_name_en', 'name_en': 'full_name_en',
            'ইমেইল': 'email', 'email': 'email',
            'ফোন': 'phone', 'phone': 'phone',
            'এনআইডি': 'nid', 'nid': 'nid',
            'শিক্ষাগত যোগ্যতা': 'education_qualification',
            'অভিজ্ঞতা (বছর)': 'years_of_experience',
            'দক্ষতার ক্ষেত্র': 'expertise_area', 'expertise_area': 'expertise_area',
            'স্ট্যাটাস': 'status', 'status': 'status',
            'অনুমোদন স্ট্যাটাস': 'approval_status', 'approval_status': 'approval_status',
        }

        results = {'created': 0, 'updated': 0, 'errors': []}
        for row_idx, row in enumerate(rows_iter, start=2):
            try:
                if isinstance(row, dict):
                    raw = row
                else:
                    raw = dict(zip(header_row, [str(c).strip() if c is not None else '' for c in row]))

                data = {}
                for k, v in raw.items():
                    mapped = field_map.get(k.strip().lower(), k.strip().lower())
                    data[mapped] = v.strip() if v else ''

                trainer_no = data.get('trainer_no', '').strip()
                if not trainer_no:
                    results['errors'].append(f'সারি {row_idx}: প্রশিক্ষক নং আবশ্যক')
                    continue

                existing = Trainer.objects.filter(trainer_no=trainer_no).first()
                if existing:
                    profile = existing.user
                    if data.get('full_name_bn'):
                        profile.full_name_bn = data['full_name_bn']
                    if data.get('full_name_en'):
                        profile.full_name_en = data['full_name_en']
                    if data.get('phone'):
                        profile.phone = data['phone']
                    if data.get('email'):
                        profile.email = data['email']
                    profile.save()
                    if data.get('education_qualification') is not None:
                        existing.education_qualification = data.get('education_qualification') or ''
                    if data.get('years_of_experience') is not None:
                        existing.years_of_experience = int(data['years_of_experience']) if data['years_of_experience'] else None
                    if data.get('expertise_area') is not None:
                        existing.expertise_area = data.get('expertise_area') or ''
                    if data.get('status') is not None:
                        existing.status = data['status']
                    if data.get('approval_status') is not None:
                        existing.approval_status = data['approval_status']
                    existing.save()
                    results['updated'] += 1
                else:
                    results['errors'].append(f'সারি {row_idx}: "{trainer_no}" প্রশিক্ষক নং পাওয়া যায়নি')
            except Exception as e:
                results['errors'].append(f'সারি {row_idx}: {str(e)}')

        return Response(results)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Template'
        headers = ['প্রশিক্ষক নং', 'নাম (বাংলা)', 'নাম (ইংরেজি)', 'ইমেইল', 'ফোন',
                   'এনআইডি', 'শিক্ষাগত যোগ্যতা', 'অভিজ্ঞতা (বছর)', 'দক্ষতার ক্ষেত্র',
                   'স্ট্যাটাস', 'অনুমোদন স্ট্যাটাস']
        ws.append(headers)
        sample = ['', 'উদাহরণ নাম', 'Example Name', 'email@example.com', '০১৭XXXXXXXX',
                  '', '', '', '', 'pending', 'pending']
        ws.append(sample)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = openpyxl.styles.Font(bold=True)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="trainer_import_template.xlsx"'
        wb.save(response)
        return response


class TrainerMappingViewSet(viewsets.ModelViewSet):
    queryset = TrainerMapping.objects.select_related(
        'trainer', 'center', 'course', 'approved_by',
    ).all()
    serializer_class = TrainerMappingSerializer
    filter_backends = (DjangoFilterBackend, filters.SearchFilter)
    filterset_fields = ('trainer', 'center', 'course', 'status', 'is_primary')
    search_fields = ('trainer__trainer_no', 'center__code', 'course__code')

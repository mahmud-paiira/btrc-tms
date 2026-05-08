from django.db.models import Count, Q, Avg
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response

from .models import JobPlacement, JobTracking
from .serializers import (
    JobPlacementSerializer,
    ReleaseJobSerializer,
    JobTrackingSerializer,
    BatchSummarySerializer,
)


class IsCenterAdminOrHeadOffice(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        return (
            request.user.user_type in ('center_admin', 'head_office')
            or request.user.is_superuser
        )


class CenterJobPlacementViewSet(viewsets.ViewSet):
    permission_classes = [permissions.IsAuthenticated, IsCenterAdminOrHeadOffice]

    def get_center_id(self, request):
        if request.user.user_type == 'head_office':
            return request.query_params.get('center')
        return request.user.center_id

    def create(self, request):
        serializer = JobPlacementSerializer(
            data=request.data, context={'request': request},
        )
        serializer.is_valid(raise_exception=True)
        serializer.save(created_by=request.user)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['put'], url_path='release')
    def release(self, request, pk=None):
        try:
            placement = JobPlacement.objects.get(pk=pk)
        except JobPlacement.DoesNotExist:
            return Response(
                {'detail': 'চাকরি স্থাপনটি খুঁজে পাওয়া যায়নি।'},
                status=status.HTTP_404_NOT_FOUND,
            )

        center_id = self.get_center_id(request)
        if center_id and str(placement.batch.center_id) != str(center_id):
            return Response(
                {'detail': 'আপনি শুধুমাত্র নিজ কেন্দ্রের ডাটা পরিবর্তন করতে পারেন।'},
                status=status.HTTP_403_FORBIDDEN,
            )

        serializer = ReleaseJobSerializer(
            placement, data=request.data, partial=True,
        )
        serializer.is_valid(raise_exception=True)

        placement.release_date = serializer.validated_data['release_date']
        placement.is_current = False
        placement.save(update_fields=['release_date', 'is_current'])

        return Response(JobPlacementSerializer(placement).data)

    @action(detail=False, methods=['get'], url_path='batch-summary/(?P<batch_id>[^/.]+)')
    def batch_summary(self, request, batch_id=None):
        from apps.trainees.models import Trainee
        from apps.batches.models import Batch

        center_id = self.get_center_id(request)
        batch = Batch.objects.filter(pk=batch_id).first()
        if not batch:
            return Response(
                {'detail': 'ব্যাচটি খুঁজে পাওয়া যায়নি।'},
                status=status.HTTP_404_NOT_FOUND,
            )
        if center_id and str(batch.center_id) != str(center_id):
            return Response(
                {'detail': 'আপনি শুধুমাত্র নিজ কেন্দ্রের ব্যাচ দেখতে পারেন।'},
                status=status.HTTP_403_FORBIDDEN,
            )

        total_trainees = Trainee.objects.filter(batch_id=batch_id).count()

        placements = JobPlacement.objects.filter(batch_id=batch_id)
        placed_count = placements.count()

        by_type = {}
        for etype, _ in JobPlacement.EmploymentType.choices:
            by_type[etype] = placements.filter(employment_type=etype).count()

        currently_employed = placements.filter(is_current=True).count()
        avg_salary = placements.aggregate(avg=Avg('salary'))['avg'] or 0

        placement_rate = round(
            (placed_count / total_trainees * 100) if total_trainees else 0, 2,
        )

        return Response({
            'batch_id': int(batch_id),
            'batch_name': batch.batch_name_bn,
            'total_trainees': total_trainees,
            'placed_count': placed_count,
            'placement_rate': placement_rate,
            'by_type': by_type,
            'currently_employed': currently_employed,
            'avg_salary': float(avg_salary),
        })

    @action(detail=False, methods=['post'], url_path='tracking')
    def add_tracking(self, request):
        serializer = JobTrackingSerializer(
            data=request.data, context={'request': request},
        )
        serializer.is_valid(raise_exception=True)
        tracking = serializer.save()
        return Response(
            JobTrackingSerializer(tracking).data,
            status=status.HTTP_201_CREATED,
        )

    @action(detail=False, methods=['get'], url_path='list')
    def list_placements(self, request):
        center_id = self.get_center_id(request)
        qs = JobPlacement.objects.select_related(
            'trainee__user', 'batch', 'created_by',
        ).all()

        if center_id:
            qs = qs.filter(batch__center_id=center_id)

        batch = request.query_params.get('batch')
        if batch:
            qs = qs.filter(batch_id=batch)

        trainee = request.query_params.get('trainee')
        if trainee:
            qs = qs.filter(trainee_id=trainee)

        current = request.query_params.get('is_current')
        if current in ('true', 'false'):
            qs = qs.filter(is_current=current == 'true')

        page = int(request.query_params.get('page', 1))
        page_size = int(request.query_params.get('page_size', 20))
        start = (page - 1) * page_size
        end = start + page_size
        total = qs.count()
        results = qs[start:end]

        return Response({
            'count': total,
            'page': page,
            'page_size': page_size,
            'results': JobPlacementSerializer(results, many=True).data,
        })

    @action(detail=False, methods=['get'], url_path='trackings/(?P<placement_id>[^/.]+)')
    def placement_trackings(self, request, placement_id=None):
        qs = JobTracking.objects.filter(
            job_placement_id=placement_id,
        ).select_related('tracked_by').order_by('tracking_month')
        return Response(JobTrackingSerializer(qs, many=True).data)

    @action(detail=False, methods=['get'], url_path='batch-summary/(?P<batch_id>[^/.]+)/export')
    def export_summary(self, request, batch_id=None):
        from apps.trainees.models import Trainee
        from apps.batches.models import Batch
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side
        from django.http import HttpResponse

        center_id = self.get_center_id(request)
        batch = Batch.objects.filter(pk=batch_id).first()
        if not batch:
            return Response({'detail': 'ব্যাচটি খুঁজে পাওয়া যায়নি।'}, status=404)
        if center_id and str(batch.center_id) != str(center_id):
            return Response({'detail': 'অননুমোদিত।'}, status=403)

        placements = JobPlacement.objects.filter(batch_id=batch_id).select_related('trainee__user')
        total_trainees = Trainee.objects.filter(batch_id=batch_id).count()

        wb = Workbook()
        ws = wb.active
        ws.title = 'Placement Summary'
        header_font = Font(bold=True, size=12)
        header_align = Alignment(horizontal='center', vertical='center')
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        ws.merge_cells('A1:H1')
        ws['A1'] = f'চাকরি স্থাপন প্রতিবেদন - {batch.batch_name_bn}'
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 30

        headers = [
            'রেজি নং', 'প্রশিক্ষণার্থীর নাম', 'কর্মসংস্থানের ধরণ',
            'নিয়োগকর্তা', 'পদবী', 'শুরুর তারিখ', 'বেতন', 'বর্তমান',
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=h)
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        row = 4
        for p in placements:
            data = [
                p.trainee.registration_no,
                p.trainee.user.full_name_bn,
                p.get_employment_type_display(),
                p.employer_name,
                p.designation_bn,
                p.start_date.isoformat() if p.start_date else '',
                float(p.salary),
                'হ্যাঁ' if p.is_current else 'না',
            ]
            for col, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = thin_border
            row += 1

        row += 1
        ws.cell(row=row, column=1, value='মোট প্রশিক্ষণার্থী').font = Font(bold=True)
        ws.cell(row=row, column=2, value=total_trainees)
        row += 1
        ws.cell(row=row, column=1, value='স্থাপিত').font = Font(bold=True)
        ws.cell(row=row, column=2, value=placements.count())
        row += 1
        rate = round((placements.count() / total_trainees * 100) if total_trainees else 0, 2)
        ws.cell(row=row, column=1, value='স্থাপনের হার (%)').font = Font(bold=True)
        ws.cell(row=row, column=2, value=rate)

        ws.column_dimensions['A'].width = 16
        ws.column_dimensions['B'].width = 25
        ws.column_dimensions['C'].width = 18
        ws.column_dimensions['D'].width = 25
        ws.column_dimensions['E'].width = 20
        ws.column_dimensions['F'].width = 14
        ws.column_dimensions['G'].width = 12
        ws.column_dimensions['H'].width = 10

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        filename = f'placement_summary_batch_{batch_id}.xlsx'
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

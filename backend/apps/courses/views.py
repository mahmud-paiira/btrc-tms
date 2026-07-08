import csv
import io
from datetime import date

import openpyxl
from django.http import HttpResponse
from rest_framework import viewsets, filters, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from .models import (
    Course,
    CourseConfiguration,
    CourseBill,
    CourseChapter,
    UnitOfCompetency,
)
from .serializers import (
    CourseListSerializer,
    CourseDetailSerializer,
    CourseWriteSerializer,
    CourseConfigurationSerializer,
    CourseBillSerializer,
    CourseChapterSerializer,
    UnitOfCompetencySerializer,
)


class CourseViewSet(viewsets.ModelViewSet):
    queryset = Course.objects.select_related(
        'configuration', 'created_by',
    ).prefetch_related(
        'bills', 'chapters', 'competencies',
    ).all()

    filter_backends = (DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter)
    filterset_fields = ('course_type', 'term', 'session', 'status', 'stipend_eligible')
    search_fields = ('code', 'name_bn', 'name_en')
    ordering_fields = ('code', 'created_at', 'fee', 'duration_months')
    ordering = ('-created_at',)

    def get_serializer_class(self):
        if self.action == 'list':
            return CourseListSerializer
        if self.action in ('create', 'update', 'partial_update'):
            return CourseWriteSerializer
        return CourseDetailSerializer

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': 'কোন আইডি প্রদান করা হয়নি'}, status=400)
        deleted = 0
        errors = []
        for pk in ids:
            try:
                obj = self.get_queryset().get(pk=pk)
                obj.delete()
                deleted += 1
            except Exception as e:
                errors.append(str(e))
        return Response({'deleted': deleted, 'errors': errors})

    @action(detail=False, methods=['get'], url_path='export-list')
    def export_list(self, request):
        qs = self.filter_queryset(self.get_queryset())
        ids = request.query_params.get('ids', '')
        if ids:
            id_list = [int(x) for x in ids.split(',') if x.isdigit()]
            if id_list:
                qs = qs.filter(id__in=id_list)
        fmt = request.query_params.get('file_format', 'xlsx')

        headers = [
            'কোড', 'নাম (বাংলা)', 'নাম (ইংরেজি)', 'ধরণ', 'মেয়াদ (মাস)',
            'ফি', 'প্রকল্প', 'স্পন্সর', 'অবস্থা', 'তৈরির তারিখ',
        ]
        rows = []
        for c in qs:
            rows.append([
                c.code, c.name_bn, c.name_en,
                c.get_course_type_display() if c.course_type else '',
                c.duration_months or '',
                str(c.fee) if c.fee else '',
                c.project_name or '', c.project_sponsor or '',
                c.get_status_display() if c.status else '',
                c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else '',
            ])

        if fmt == 'xlsx':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Courses'
            ws.append(headers)
            for row in rows:
                ws.append(row)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="courses_{date.today().isoformat()}.xlsx"'
            return response
        else:
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)
            response = HttpResponse(
                output.getvalue().encode('utf-8-sig'),
                content_type='text/csv',
            )
            response['Content-Disposition'] = f'attachment; filename="courses_{date.today().isoformat()}.csv"'
            return response

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def import_list(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'ফাইল নির্বাচন করুন'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            rows_iter = iter(ws.iter_rows(values_only=True))
            header_row = [str(c).strip().lower() if c is not None else '' for c in next(rows_iter)]
        except Exception:
            file.seek(0)
            try:
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                header_row = [h.strip().lower() for h in reader.fieldnames]
                rows_iter = reader
            except Exception:
                return Response({'error': 'ভুল ফাইল ফরম্যাট। Excel (.xlsx) বা CSV ফাইল আপলোড করুন।'}, status=400)

        bn_required = {'কোড'}
        header_set = set(header_row)
        if not bn_required.issubset(header_set):
            return Response({
                'error': 'প্রয়োজনীয় কলাম নেই। হেডারে "কোড" থাকা আবশ্যক।',
                'detected_headers': header_row,
            }, status=400)

        field_map = {
            'কোড': 'code', 'code': 'code',
            'নাম (বাংলা)': 'name_bn', 'name_bn': 'name_bn',
            'নাম (ইংরেজি)': 'name_en', 'name_en': 'name_en',
            'ধরণ': 'course_type', 'course_type': 'course_type',
            'মেয়াদ (মাস)': 'duration_months', 'duration_months': 'duration_months',
            'ফি': 'fee', 'fee': 'fee',
            'প্রকল্প': 'project_name', 'project_name': 'project_name',
            'স্পন্সর': 'project_sponsor', 'project_sponsor': 'project_sponsor',
            'অবস্থা': 'status', 'status': 'status',
        }

        results = {'created': 0, 'updated': 0, 'errors': []}
        for row_idx, row in enumerate(rows_iter, start=2):
            try:
                if isinstance(row, dict):
                    raw = row
                else:
                    raw = dict(zip(header_row, [str(c).strip() if c is not None else '' for c in row]))

                data = {}
                for k, v in raw.items():
                    mapped = field_map.get(k.strip().lower(), k.strip().lower())
                    data[mapped] = v.strip() if v else ''

                code = data.get('code', '').strip()
                if not code:
                    results['errors'].append(f'সারি {row_idx}: কোড আবশ্যক')
                    continue

                existing = Course.objects.filter(code=code).first()
                if existing:
                    if data.get('name_bn'):
                        existing.name_bn = data['name_bn']
                    if data.get('name_en'):
                        existing.name_en = data['name_en']
                    if data.get('course_type'):
                        existing.course_type = data['course_type']
                    if data.get('duration_months') is not None:
                        try:
                            existing.duration_months = int(data['duration_months']) if data['duration_months'] else None
                        except ValueError:
                            pass
                    if data.get('fee') is not None:
                        try:
                            existing.fee = float(data['fee']) if data['fee'] else None
                        except ValueError:
                            pass
                    if data.get('project_name') is not None:
                        existing.project_name = data.get('project_name') or ''
                    if data.get('project_sponsor') is not None:
                        existing.project_sponsor = data.get('project_sponsor') or ''
                    if data.get('status') is not None:
                        existing.status = data['status']
                    existing.save()
                    results['updated'] += 1
                else:
                    results['errors'].append(f'সারি {row_idx}: "{code}" কোডের কোর্স পাওয়া যায়নি')
            except Exception as e:
                results['errors'].append(f'সারি {row_idx}: {str(e)}')

        return Response(results)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Template'
        headers = ['কোড', 'নাম (বাংলা)', 'নাম (ইংরেজি)', 'ধরণ', 'মেয়াদ (মাস)',
                   'ফি', 'প্রকল্প', 'স্পন্সর', 'অবস্থা']
        ws.append(headers)
        sample = ['', 'উদাহরণ কোর্স', 'Example Course', 'vocational', '6',
                  '৫০০০', '', '', 'active']
        ws.append(sample)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = openpyxl.styles.Font(bold=True)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="course_import_template.xlsx"'
        wb.save(response)
        return response


class CourseConfigurationViewSet(viewsets.ModelViewSet):
    queryset = CourseConfiguration.objects.select_related('course').all()
    serializer_class = CourseConfigurationSerializer
    filterset_fields = ('course',)


class CourseBillViewSet(viewsets.ModelViewSet):
    queryset = CourseBill.objects.select_related('course').all()
    serializer_class = CourseBillSerializer
    filterset_fields = ('course', 'is_mandatory')


class CourseChapterViewSet(viewsets.ModelViewSet):
    queryset = CourseChapter.objects.select_related('course').all()
    serializer_class = CourseChapterSerializer
    filterset_fields = ('course',)
    ordering = ('course', 'chapter_no')


class UnitOfCompetencyViewSet(viewsets.ModelViewSet):
    queryset = UnitOfCompetency.objects.select_related('course').all()
    serializer_class = UnitOfCompetencySerializer
    filterset_fields = ('course',)
    search_fields = ('code', 'name_bn', 'name_en')

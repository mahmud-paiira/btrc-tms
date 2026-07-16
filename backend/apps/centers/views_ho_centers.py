import csv
import io
from datetime import date, timedelta

import openpyxl
from django.db.models import Count, Avg, Sum, Q
from django.db.models.functions import TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import viewsets, status, permissions, filters as drf_filters
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend

from .models import Center, Infrastructure, Employee, ActionLog
from apps.accounts.models import User
from .serializers import (
    CenterListSerializer, CenterDetailSerializer,
    InfrastructureSerializer, EmployeeSerializer,
)
from apps.trainees.models import Trainee
from apps.batches.models import Batch
from apps.attendance.models import AttendanceSummary
from apps.jobplacement.models import JobPlacement
from apps.trainers.models import Trainer
from apps.assessors.models import Assessor


class IsHeadOffice(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        return request.user.user_type == 'head_office' or request.user.is_superuser


class HOCenterPagination(PageNumberPagination):
    page_size = 25
    page_size_query_param = 'page_size'
    max_page_size = 999


class HOCenterViewSet(viewsets.ModelViewSet):
    queryset = Center.objects.prefetch_related('infrastructures', 'employees').all()
    permission_classes = [permissions.IsAuthenticated, IsHeadOffice]
    lookup_value_regex = '\d+'
    pagination_class = HOCenterPagination
    filter_backends = (DjangoFilterBackend, drf_filters.SearchFilter, drf_filters.OrderingFilter)
    filterset_fields = ('status',)
    search_fields = ('code', 'name_bn', 'name_en', 'phone', 'email', 'address')
    ordering_fields = ('code', 'name_en', 'created_at')
    ordering = ('name_en',)

    def get_serializer_class(self):
        if self.action == 'list':
            return CenterListSerializer
        return CenterDetailSerializer

    def perform_create(self, serializer):
        center = serializer.save()
        ActionLog.objects.create(
            user=self.request.user,
            action='created center',
            target_type='Center',
            target_id=str(center.id),
            description=f'Created center {center.code} - {center.name_bn}',
            ip_address=self.request.META.get('REMOTE_ADDR', ''),
        )

        email = f'center{center.code}@brtc.gov.bd'
        if not User.objects.filter(email=email).exists():
            User.objects.create_user(
                email=email,
                password='center@123',
                user_type='center_admin',
                center=center,
                full_name_bn=f'কেন্দ্র প্রশাসক - {center.name_bn}',
                full_name_en=f'Center Admin - {center.name_en}',
                nid=f'{center.code.zfill(8)}NID00000',
                phone=f'019{center.code.zfill(8)}',
                birth_certificate_no=f'{center.code.zfill(10)}000000',
                is_active=True,
            )

    def perform_update(self, serializer):
        center = serializer.save()
        ActionLog.objects.create(
            user=self.request.user,
            action='updated center',
            target_type='Center',
            target_id=str(center.id),
            description=f'Updated center {center.code} - {center.name_bn}',
            ip_address=self.request.META.get('REMOTE_ADDR', ''),
        )

    def perform_destroy(self, instance):
        if instance.batches.exists():
            from rest_framework.exceptions import ValidationError
            raise ValidationError('এই কেন্দ্রে ব্যাচ আছে। প্রথমে ব্যাচ স্থানান্তর বা মুছুন।')
        ActionLog.objects.create(
            user=self.request.user,
            action='deleted center',
            target_type='Center',
            target_id=str(instance.id),
            description=f'Deleted center {instance.code} - {instance.name_bn}',
            ip_address=self.request.META.get('REMOTE_ADDR', ''),
        )
        instance.delete()

    @action(detail=False, methods=['get'], url_path='export-list')
    def export(self, request):
        qs = self.filter_queryset(self.get_queryset())
        ids = request.query_params.get('ids', '')
        if ids:
            id_list = [int(x) for x in ids.split(',') if x.isdigit()]
            if id_list:
                qs = qs.filter(id__in=id_list)
        fmt = request.query_params.get('file_format', 'xlsx')

        headers = [
            'কোড', 'নাম (বাংলা)', 'নাম (ইংরেজি)', 'সংক্ষিপ্ত নাম',
            'ফোন', 'ইমেইল', 'ওয়েবসাইট', 'ঠিকানা',
            'যোগাযোগ ব্যক্তি', 'যোগাযোগ মোবাইল', 'স্ট্যাটাস', 'তৈরির তারিখ',
        ]
        rows = []
        for c in qs:
            rows.append([
                c.code, c.name_bn, c.name_en, c.short_name_bn,
                c.phone, c.email, c.website_url, c.address,
                c.contact_person_name, c.contact_person_phone,
                'সক্রিয়' if c.status == 'active' else 'স্থগিত',
                c.created_at.strftime('%Y-%m-%d %H:%M') if c.created_at else '',
            ])
        rows.sort(key=lambda r: int(r[0]) if r[0] and r[0].isdigit() else 9999)

        if fmt == 'xlsx':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Centers'
            ws.append(headers)
            for row in rows:
                ws.append(row)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="centers_{date.today().isoformat()}.xlsx"'
            return response
        else:
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)
            response = HttpResponse(
                output.getvalue().encode('utf-8-sig'),
                content_type='text/csv',
            )
            response['Content-Disposition'] = f'attachment; filename="centers_{date.today().isoformat()}.csv"'
            return response

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def import_centers(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'ফাইল নির্বাচন করুন'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            rows_iter = iter(ws.iter_rows(values_only=True))
            header_row = [str(c).strip().lower() if c is not None else '' for c in next(rows_iter)]
        except Exception:
            file.seek(0)
            try:
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                header_row = [h.strip().lower() for h in reader.fieldnames]
                rows_iter = reader
            except Exception:
                return Response({'error': 'ভুল ফাইল ফরম্যাট। Excel (.xlsx) বা CSV ফাইল আপলোড করুন।'}, status=400)

        bn_required = {'নাম (বাংলা)', 'নাম (ইংরেজি)'}
        en_required = {'name_bn', 'name_en'}
        header_set = set(header_row)
        if not bn_required.issubset(header_set) and not en_required.issubset(header_set):
            return Response({
                'error': f'প্রয়োজনীয় কলাম নেই। হেডার হতে হবে: {", ".join(sorted(bn_required))}',
                'detected_headers': header_row,
            }, status=400)

        results = {'created': 0, 'updated': 0, 'center_admins_created': 0, 'errors': []}
        field_map = {
            'name_bn': 'name_bn', 'নাম (বাংলা)': 'name_bn',
            'name_en': 'name_en', 'নাম (ইংরেজি)': 'name_en',
            'short_name_bn': 'short_name_bn', 'সংক্ষিপ্ত নাম': 'short_name_bn',
            'phone': 'phone', 'ফোন': 'phone',
            'email': 'email', 'ইমেইল': 'email',
            'website_url': 'website_url', 'ওয়েবসাইট': 'website_url',
            'address': 'address', 'ঠিকানা': 'address',
            'contact_person_name': 'contact_person_name', 'যোগাযোগ ব্যক্তি': 'contact_person_name',
            'contact_person_phone': 'contact_person_phone', 'যোগাযোগ মোবাইল': 'contact_person_phone',
            'code': 'code', 'কোড': 'code',
            'স্ট্যাটাস': 'status', 'status': 'status',
        }

        for row_idx, row in enumerate(rows_iter, start=2):
            try:
                if isinstance(row, dict):
                    raw = row
                else:
                    raw = dict(zip(header_row, [str(c).strip() if c is not None else '' for c in row]))

                data = {}
                for k, v in raw.items():
                    mapped = field_map.get(k.strip().lower(), k.strip().lower())
                    data[mapped] = v.strip() if v else ''

                name_bn = data.get('name_bn', '').strip()
                name_en = data.get('name_en', '').strip()
                if not name_bn or not name_en:
                    results['errors'].append(f'সারি {row_idx}: নাম (বাংলা) এবং নাম (ইংরেজি) আবশ্যক')
                    continue

                code = data.get('code', '').strip()
                existing = None
                if code:
                    existing = Center.objects.filter(code=code).first()
                if not existing:
                    existing = Center.objects.filter(name_bn=name_bn).first()

                if existing:
                    for fld in ['name_en', 'short_name_bn', 'phone', 'email', 'website_url', 'address', 'contact_person_name', 'contact_person_phone']:
                        if data.get(fld):
                            setattr(existing, fld, data[fld])
                    existing.save()
                    results['updated'] += 1
                else:
                    cleaned = {}
                    for k, v in data.items():
                        if k in [f.name for f in Center._meta.fields if f.editable] and v:
                            cleaned[k] = v
                    if 'status' not in cleaned:
                        cleaned['status'] = 'active'
                    center = Center.objects.create(**cleaned)
                    results['created'] += 1

                    email = f'center{center.code}@brtc.gov.bd'
                    if not User.objects.filter(email=email).exists():
                        try:
                            User.objects.create_user(
                                email=email,
                                password='center@123',
                                user_type='center_admin',
                                center=center,
                                full_name_bn=f'কেন্দ্র প্রশাসক - {center.name_bn}',
                                full_name_en=f'Center Admin - {center.name_en}',
                                nid=f'{center.code.zfill(8)}NID00000',
                                phone=f'019{center.code.zfill(8)}',
                                birth_certificate_no=f'{center.code.zfill(10)}000000',
                                is_active=True,
                            )
                            results['center_admins_created'] += 1
                        except Exception as e:
                            results['errors'].append(f'সারি {row_idx}: কেন্দ্র প্রশাসক তৈরি ব্যর্থ ({email}: {e})')
            except Exception as e:
                results['errors'].append(f'সারি {row_idx}: {str(e)}')

        return Response(results)

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Template'
        headers = ['কোড', 'নাম (বাংলা)', 'নাম (ইংরেজি)', 'সংক্ষিপ্ত নাম',
                   'ফোন', 'ইমেইল', 'ওয়েবসাইট', 'ঠিকানা',
                   'যোগাযোগ ব্যক্তি', 'যোগাযোগ মোবাইল', 'স্ট্যাটাস']
        ws.append(headers)
        sample = ['', 'উদাহরণ কেন্দ্র', 'Example Center', '', '০১৭XXXXXXXX', '', '', 'ঢাকা', '', '', 'active']
        ws.append(sample)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = openpyxl.styles.Font(bold=True)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="center_import_template.xlsx"'
        wb.save(response)
        return response

    @action(detail=True, methods=['post'])
    def toggle(self, request, pk=None):
        center = self.get_object()
        center.status = Center.Status.SUSPENDED if center.status == Center.Status.ACTIVE else Center.Status.ACTIVE
        center.save(update_fields=['status'])
        ActionLog.objects.create(
            user=request.user,
            action='toggled center status',
            target_type='Center',
            target_id=str(center.id),
            description=f'Set center {center.code} to {center.status}',
            ip_address=request.META.get('REMOTE_ADDR', ''),
        )
        return Response({'status': center.status})

    @action(detail=True, methods=['get'])
    def stats(self, request, pk=None):
        center = self.get_object()
        return Response({
            'center_name': center.name_bn,
            'center_code': center.code,
            'trainee_count': center.total_trainees,
            'active_batch_count': center.active_batches,
            'attendance_rate': float(center.attendance_rate),
            'placement_rate': float(center.placement_rate),
            'total_batches': center.total_batches,
            'running_batches': center.running_batches,
            'completed_batches': center.completed_batches,
            'enrolled_trainees': center.enrolled_trainees,
            'completed_trainees': center.completed_trainees,
            'dropped_trainees': center.dropped_trainees,
            'monthly_enrollment': [],
        })

    @action(detail=True, methods=['get', 'post'])
    def infrastructure(self, request, pk=None):
        center = self.get_object()
        if request.method == 'GET':
            qs = center.infrastructures.all()
            return Response(InfrastructureSerializer(qs, many=True).data)
        serializer = InfrastructureSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(center=center)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=True, methods=['get', 'post'])
    def employees(self, request, pk=None):
        center = self.get_object()
        if request.method == 'GET':
            qs = center.employees.select_related('user').all()
            return Response(EmployeeSerializer(qs, many=True).data)
        serializer = EmployeeSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(center=center)
        return Response(serializer.data, status=status.HTTP_201_CREATED)

    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': 'কোন আইডি প্রদান করা হয়নি'}, status=400)
        deleted = 0
        errors = []
        for pk in ids:
            try:
                obj = self.get_queryset().get(pk=pk)
                self.perform_destroy(obj)
                deleted += 1
            except Exception as e:
                msg = str(e.detail[0]) if hasattr(e, 'detail') and isinstance(e.detail, list) else str(e)
                errors.append(msg)
        return Response({'deleted': deleted, 'errors': errors})

class HOInfrastructureViewSet(viewsets.ModelViewSet):
    queryset = Infrastructure.objects.select_related('center').all()
    serializer_class = InfrastructureSerializer
    permission_classes = [permissions.IsAuthenticated, IsHeadOffice]
    filter_backends = (DjangoFilterBackend, drf_filters.SearchFilter)
    filterset_fields = ('center', 'status')
    search_fields = ('room_no', 'location_bn', 'location_en')


class HOEmployeeViewSet(viewsets.ModelViewSet):
    queryset = Employee.objects.select_related('user', 'center').all()
    serializer_class = EmployeeSerializer
    permission_classes = [permissions.IsAuthenticated, IsHeadOffice]
    filter_backends = (DjangoFilterBackend, drf_filters.SearchFilter)
    filterset_fields = ('center', 'status', 'is_contact_person')
    search_fields = ('employee_no', 'user__full_name_bn', 'user__full_name_en', 'designation_bn')

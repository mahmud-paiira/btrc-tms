import csv
import io
from datetime import date

import openpyxl
from django.http import HttpResponse
from rest_framework import viewsets, filters, permissions, status
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend

from .models import Trainee
from .serializers import (
    TraineeListSerializer,
    TraineeDetailSerializer,
    TraineeWriteSerializer,
)


class IsCenterAdmin(permissions.BasePermission):
    def has_permission(self, request, view):
        return request.user.user_type == 'center_admin' or request.user.is_superuser


class TraineeViewSet(viewsets.ModelViewSet):
    queryset = Trainee.objects.select_related(
        'user', 'center', 'batch'
    ).all()
    permission_classes = [IsAuthenticated, IsCenterAdmin]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ('center', 'batch', 'status')
    search_fields = (
        'registration_no', 'user__full_name_bn', 'user__full_name_en',
        'user__email', 'user__phone', 'user__nid',
    )
    ordering_fields = ('enrollment_date', 'registration_no', 'status')
    ordering = ('-enrollment_date',)

    def get_serializer_class(self):
        if self.action == 'list':
            return TraineeListSerializer
        elif self.action in ('create', 'update', 'partial_update'):
            return TraineeWriteSerializer
        return TraineeDetailSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        user = self.request.user
        if user.user_type == 'center_admin' and user.center:
            qs = qs.filter(center=user.center)
        return qs

    @action(detail=False, methods=['get'], url_path='export-list')
    def export_list(self, request):
        qs = self.filter_queryset(self.get_queryset())
        ids = request.query_params.get('ids', '')
        if ids:
            id_list = [int(x) for x in ids.split(',') if x.isdigit()]
            if id_list:
                qs = qs.filter(id__in=id_list)
        fmt = request.query_params.get('file_format', 'xlsx')

        headers = [
            'রেজি. নং', 'নাম (বাংলা)', 'নাম (ইংরেজি)', 'ইমেইল', 'ফোন',
            'এনআইডি', 'কেন্দ্রের কোড', 'কেন্দ্র', 'ব্যাচ', 'অবস্থা', 'নথিভুক্তির তারিখ',
        ]
        rows = []
        for t in qs:
            rows.append([
                t.registration_no, t.user.full_name_bn, t.user.full_name_en,
                t.user.email, t.user.phone, t.user.nid or '',
                t.center.code if t.center else '',
                t.center.name_bn if t.center else '',
                t.batch.batch_name_bn if t.batch else '',
                t.get_status_display() if t.status else '',
                t.enrollment_date.strftime('%Y-%m-%d') if t.enrollment_date else '',
            ])

        if fmt == 'xlsx':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Trainees'
            ws.append(headers)
            for row in rows:
                ws.append(row)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="trainees_{date.today().isoformat()}.xlsx"'
            return response
        else:
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)
            response = HttpResponse(
                output.getvalue().encode('utf-8-sig'),
                content_type='text/csv',
            )
            response['Content-Disposition'] = f'attachment; filename="trainees_{date.today().isoformat()}.csv"'
            return response

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def import_list(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'ফাইল নির্বাচন করুন'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            rows_iter = iter(ws.iter_rows(values_only=True))
            header_row = [str(c).strip().lower() if c is not None else '' for c in next(rows_iter)]
        except Exception:
            file.seek(0)
            try:
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                header_row = [h.strip().lower() for h in reader.fieldnames]
                rows_iter = reader
            except Exception:
                return Response({'error': 'ভুল ফাইল ফরম্যাট। Excel (.xlsx) বা CSV ফাইল আপলোড করুন।'}, status=400)

        bn_required = {'রেজি. নং'}
        header_set = set(header_row)
        if not bn_required.issubset(header_set):
            return Response({
                'error': 'প্রয়োজনীয় কলাম নেই। হেডারে "রেজি. নং" থাকা আবশ্যক।',
                'detected_headers': header_row,
            }, status=400)

        field_map = {
            'রেজি. নং': 'registration_no', 'registration_no': 'registration_no',
            'নাম (বাংলা)': 'full_name_bn', 'name_bn': 'full_name_bn',
            'নাম (ইংরেজি)': 'full_name_en', 'name_en': 'full_name_en',
            'ইমেইল': 'email', 'email': 'email',
            'ফোন': 'phone', 'phone': 'phone',
            'এনআইডি': 'nid', 'nid': 'nid',
            'অবস্থা': 'status', 'status': 'status',
            'কেন্দ্রের কোড': 'center_code', 'center_code': 'center_code',
            'কেন্দ্রের নাম': 'center_name', 'center_name': 'center_name',
        }

        results = {'created': 0, 'updated': 0, 'errors': []}
        for row_idx, row in enumerate(rows_iter, start=2):
            try:
                if isinstance(row, dict):
                    raw = row
                else:
                    raw = dict(zip(header_row, [str(c).strip() if c is not None else '' for c in row]))

                data = {}
                for k, v in raw.items():
                    mapped = field_map.get(k.strip().lower(), k.strip().lower())
                    data[mapped] = v.strip() if v else ''

                reg_no = data.get('registration_no', '').strip()
                if not reg_no:
                    results['errors'].append(f'সারি {row_idx}: রেজি. নং আবশ্যক')
                    continue

                existing = Trainee.objects.filter(registration_no=reg_no).first()
                if existing:
                    if data.get('status') is not None:
                        existing.status = data['status']
                    existing.save()

                    from apps.centers.models import Center
                    center_code = data.get('center_code', '').strip()
                    center_name = data.get('center_name', '').strip()
                    if center_code:
                        center = Center.objects.filter(code__iexact=center_code).first()
                        if not center and center_name:
                            center = Center.objects.filter(name_bn=center_name).first()
                        if center:
                            existing.center = center
                            existing.save(update_fields=['center'])

                    results['updated'] += 1
                else:
                    results['errors'].append(f'সারি {row_idx}: "{reg_no}" রেজি. নং পাওয়া যায়নি')
            except Exception as e:
                results['errors'].append(f'সারি {row_idx}: {str(e)}')

        return Response(results)

    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': 'কোন আইডি প্রদান করা হয়নি'}, status=400)
        deleted = 0
        errors = []
        for pk in ids:
            try:
                obj = self.get_queryset().get(pk=pk)
                obj.delete()
                deleted += 1
            except Exception as e:
                msg = str(e.detail[0]) if hasattr(e, 'detail') and isinstance(e.detail, list) else str(e)
                errors.append(msg)
        return Response({'deleted': deleted, 'errors': errors})

    @action(detail=False, methods=['get'])
    def download_template(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Template'
        headers = ['রেজি. নং', 'নাম (বাংলা)', 'নাম (ইংরেজি)', 'ইমেইল', 'ফোন',
                   'এনআইডি', 'অবস্থা', 'কেন্দ্রের কোড', 'কেন্দ্রের নাম']
        ws.append(headers)
        sample = ['', 'উদাহরণ নাম', 'Example Name', 'email@example.com', '০১৭XXXXXXXX',
                  '', 'enrolled', 'RSH_TCU', 'রাজশাহী ট্রেনিং সেন্টার']
        ws.append(sample)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = openpyxl.styles.Font(bold=True)
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="trainee_import_template.xlsx"'
        wb.save(response)
        return response

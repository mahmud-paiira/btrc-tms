import io
import logging
import csv
import openpyxl

from datetime import date, timedelta
from django.db.models import Count, Q, Avg, F
from django.utils import timezone
from rest_framework import viewsets, status, permissions, filters as drf_filters
from rest_framework.decorators import action
from rest_framework.parsers import MultiPartParser, FormParser
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.utils.dateparse import parse_date

from .models import Circular
from .serializers import (
    CircularListSerializer, CircularDetailSerializer, CircularWriteSerializer,
)
from apps.applications.models import Application
from apps.applications.serializers import ApplicationDetailSerializer as ApplicationSerializer
from apps.centers.models import ActionLog
from django.conf import settings
from django.template.loader import render_to_string
from django.http import HttpResponse
from weasyprint import HTML

logger = logging.getLogger(__name__)


class IsHeadOffice(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        return request.user.user_type == 'head_office' or request.user.is_superuser


class HOCircularViewSet(viewsets.ModelViewSet):
    queryset = Circular.objects.prefetch_related(
        'eligible_centers', 'checklist_items',
    ).select_related('course', 'created_by').all()
    permission_classes = [permissions.IsAuthenticated, IsHeadOffice]
    filter_backends = (DjangoFilterBackend, drf_filters.SearchFilter, drf_filters.OrderingFilter)
    filterset_fields = ('eligible_centers', 'course', 'status')
    search_fields = ('title_bn', 'title_en')
    ordering_fields = ('created_at', 'application_start_date', 'total_seats', 'status')
    ordering = ('-created_at',)

    def get_serializer_class(self):
        if self.action == 'list':
            return CircularListSerializer
        if self.action in ('create', 'update', 'partial_update'):
            return CircularWriteSerializer
        return CircularDetailSerializer

    def _log(self, request, action_desc, target_id=None):
        ActionLog.objects.create(
            user=request.user,
            action=action_desc[:255],
            target_type='Circular',
            target_id=str(target_id) if target_id else '',
            description=action_desc,
            ip_address=request.META.get('REMOTE_ADDR', ''),
        )

    def perform_create(self, serializer):
        obj = serializer.save(created_by=self.request.user)
        self._log(self.request, f'Created circular {obj.title_bn}', obj.id)

    def perform_update(self, serializer):
        obj = serializer.save()
        self._log(self.request, f'Updated circular {obj.title_bn}', obj.id)

    def perform_destroy(self, instance):
        self._log(self.request, f'Deleted circular {instance.title_bn}', instance.id)
        instance.delete()

    @action(detail=False, methods=['post'])
    def bulk_delete(self, request):
        ids = request.data.get('ids', [])
        if not ids:
            return Response({'error': 'কোন আইডি প্রদান করা হয়নি'}, status=400)
        deleted = 0
        errors = []
        for pk in ids:
            try:
                obj = self.get_queryset().get(pk=pk)
                self.perform_destroy(obj)
                deleted += 1
            except Exception as e:
                errors.append(str(e))
        return Response({'deleted': deleted, 'errors': errors})

    @action(detail=True, methods=['post'])
    def publish(self, request, pk=None):
        circular = self.get_object()
        if circular.status != Circular.Status.DRAFT:
            return Response({'error': 'শুধুমাত্র খসড়া অবস্থায় প্রকাশ করা যাবে'}, status=400)
        from django.utils import timezone as tz
        today = tz.localdate()
        if circular.application_start_date and circular.application_start_date > today:
            circular.application_start_date = today
        circular.status = Circular.Status.PUBLISHED
        circular.published_at = tz.now()
        if not circular.public_url:
            import uuid
            circular.public_url = uuid.uuid4().hex[:12]
        circular.save(update_fields=['status', 'published_at', 'public_url', 'application_start_date'])
        self._log(request, f'Published circular {circular.title_bn}', circular.id)
        return Response(CircularDetailSerializer(circular).data)

    @action(detail=True, methods=['post'])
    def close(self, request, pk=None):
        circular = self.get_object()
        if circular.status != Circular.Status.PUBLISHED:
            return Response({'error': 'শুধুমাত্র প্রকাশিত সার্কুলার বন্ধ করা যাবে'}, status=400)
        circular.status = Circular.Status.CLOSED
        circular.save(update_fields=['status'])
        self._log(request, f'Closed circular {circular.title_bn}', circular.id)
        return Response(CircularDetailSerializer(circular).data)

    @action(detail=True, methods=['post'])
    def unpublish(self, request, pk=None):
        circular = self.get_object()
        if circular.status != Circular.Status.PUBLISHED:
            return Response({'error': 'শুধুমাত্র প্রকাশিত সার্কুলার খসড়ায় ফেরত নেওয়া যাবে'}, status=400)
        circular.status = Circular.Status.DRAFT
        circular.published_at = None
        circular.save(update_fields=['status', 'published_at'])
        self._log(request, f'Unpublished circular {circular.title_bn}', circular.id)
        return Response(CircularDetailSerializer(circular).data)

    @action(detail=True, methods=['post'])
    def extend(self, request, pk=None):
        circular = self.get_object()
        days = request.data.get('days', 7)
        if circular.status != Circular.Status.PUBLISHED:
            return Response({'error': 'শুধুমাত্র প্রকাশিত সার্কুলারের মেয়াদ বাড়ানো যাবে'}, status=400)
        try:
            days = int(days)
        except (TypeError, ValueError):
            return Response({'error': 'days must be an integer'}, status=400)
        circular.application_end_date += timedelta(days=days)
        circular.save(update_fields=['application_end_date'])
        self._log(request, f'Extended circular {circular.title_bn} by {days} days', circular.id)
        return Response(CircularDetailSerializer(circular).data)

    @action(detail=True, methods=['get'])
    def analytics(self, request, pk=None):
        circular = self.get_object()
        apps = Application.objects.filter(circular=circular)
        total = apps.count()
        reviewed = apps.exclude(reviewed_at__isnull=True)

        applications_by_gender = {}
        for app in apps.select_related('user__profile'):
            try:
                g = app.user.profile.gender if app.user and hasattr(app.user, 'profile') else 'unknown'
            except Exception:
                g = 'unknown'
            applications_by_gender[g] = applications_by_gender.get(g, 0) + 1

        from datetime import date
        today = date.today()
        age_groups = {'18-25': 0, '26-35': 0, '36-45': 0, '46+': 0}
        for app in apps.only('date_of_birth'):
            if app.date_of_birth:
                age = today.year - app.date_of_birth.year - (
                    (today.month, today.day) < (app.date_of_birth.month, app.date_of_birth.day)
                )
                if age <= 25:
                    age_groups['18-25'] += 1
                elif age <= 35:
                    age_groups['26-35'] += 1
                elif age <= 45:
                    age_groups['36-45'] += 1
                else:
                    age_groups['46+'] += 1

        applications_by_district = {}
        for app in apps.only('permanent_address'):
            addr = (app.permanent_address or '').strip()
            if addr:
                parts = [p.strip() for p in addr.split(',')]
                district = parts[-1] if len(parts) > 1 else addr
                applications_by_district[district] = applications_by_district.get(district, 0) + 1

        selected = apps.filter(status=Application.ApplicationStatus.SELECTED).count()
        selection_rate = round((selected / total * 100), 1) if total else 0

        avg_response_time = None
        if reviewed.count() > 0:
            delta = reviewed.annotate(
                resp_time=Avg(F('reviewed_at') - F('applied_at'))
            ).aggregate(avg=Avg('resp_time'))['avg']
            if delta:
                avg_response_time = round(delta.total_seconds() / 3600, 1)

        status_counts = apps.values('status').annotate(count=Count('id'))

        return Response({
            'applications_received': total,
            'applications_by_gender': applications_by_gender,
            'applications_by_age_group': age_groups,
            'applications_by_district': dict(sorted(
                applications_by_district.items(), key=lambda x: x[1], reverse=True,
            )[:15]),
            'selection_rate': selection_rate,
            'average_response_time_hours': avg_response_time,
            'status_breakdown': {s['status']: s['count'] for s in status_counts},
        })

    @action(detail=True, methods=['get'])
    def applications(self, request, pk=None):
        circular = self.get_object()
        qs = Application.objects.filter(circular=circular).select_related(
            'user', 'reviewed_by',
        ).order_by('-applied_at')

        status_filter = request.query_params.get('status')
        if status_filter:
            qs = qs.filter(status=status_filter)

        search_q = request.query_params.get('search', '')
        if search_q:
            qs = qs.filter(
                Q(name_bn__icontains=search_q) |
                Q(name_en__icontains=search_q) |
                Q(application_no__icontains=search_q) |
                Q(nid__icontains=search_q) |
                Q(phone__icontains=search_q)
            )

        page = self.paginate_queryset(qs)
        if page is not None:
            serializer = ApplicationSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = ApplicationSerializer(qs, many=True)
        return Response(serializer.data)

    @action(detail=True, methods=['post'])
    def approve_application(self, request, pk=None):
        circular = self.get_object()
        app_id = request.data.get('application_id')
        if not app_id:
            return Response({'error': 'application_id required'}, status=400)
        try:
            app = Application.objects.get(id=app_id, circular=circular)
        except Application.DoesNotExist:
            return Response({'error': 'আবেদন পাওয়া যায়নি'}, status=404)
        app.status = Application.ApplicationStatus.SELECTED
        app.reviewed_by = request.user
        app.reviewed_at = timezone.now()
        app.committee_decision_by = request.user
        app.committee_decision_at = timezone.now()
        app.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'committee_decision_by', 'committee_decision_at'])
        if circular.remaining_seats > 0:
            circular.remaining_seats -= 1
            circular.save(update_fields=['remaining_seats'])

        # Create user account for selected applicant
        try:
            from apps.accounts.models import User
            email = app.email or f'trainee{app.application_no.lower().replace("-", "")}@brtc.gov.bd'
            username = email.split('@')[0][:30]
            user, created = User.objects.get_or_create(
                nid=app.nid,
                defaults={
                    'email': email,
                    'username': username,
                    'full_name_bn': app.name_bn,
                    'full_name_en': app.name_en or app.name_bn,
                    'phone': app.phone,
                    'user_type': User.UserType.TRAINEE,
                    'center': app.chosen_center,
                    'is_active': True,
                },
            )
            if created:
                user.set_password('changeme')
                user.save(update_fields=['password'])
                logger.info(f'Created user account for {app.application_no} (nid={app.nid})')
            app.user = user
            app.save(update_fields=['user'])
        except Exception as e:
            logger.error(f'Failed to create user for {app.application_no}: {e}')

        self._log(request, f'Approved application {app.application_no}', app.id)
        return Response(ApplicationSerializer(app).data)

    @action(detail=True, methods=['post'])
    def reject_application(self, request, pk=None):
        app_id = request.data.get('application_id')
        remarks = request.data.get('remarks', '')
        if not app_id:
            return Response({'error': 'application_id required'}, status=400)
        try:
            app = Application.objects.get(id=app_id, circular=self.get_object())
        except Application.DoesNotExist:
            return Response({'error': 'আবেদন পাওয়া যায়নি'}, status=404)
        app.status = Application.ApplicationStatus.REJECTED
        app.reviewed_by = request.user
        app.reviewed_at = timezone.now()
        app.remarks = remarks
        app.committee_decision_by = request.user
        app.committee_decision_at = timezone.now()
        app.save(update_fields=['status', 'reviewed_by', 'reviewed_at', 'remarks', 'committee_decision_by', 'committee_decision_at'])
        self._log(request, f'Rejected application {app.application_no}', app.id)
        return Response(ApplicationSerializer(app).data)

    @action(detail=True, methods=['post'])
    def enroll_application(self, request, pk=None):
        circular = self.get_object()
        app_id = request.data.get('application_id')
        if not app_id:
            return Response({'error': 'application_id required'}, status=400)
        try:
            app = Application.objects.get(id=app_id, circular=circular)
        except Application.DoesNotExist:
            return Response({'error': 'আবেদন পাওয়া যায়নি'}, status=404)
        if app.status != Application.ApplicationStatus.SELECTED:
            return Response({'error': 'শুধুমাত্র নির্বাচিত আবেদন নথিভুক্ত করা যাবে'}, status=400)
        app.status = Application.ApplicationStatus.ENROLLED
        app.save(update_fields=['status'])
        from apps.trainees.models import Trainee
        Trainee.objects.get_or_create(
            user=app.user,
            defaults={
                'name_bn': app.name_bn, 'name_en': app.name_en or app.name_bn,
                'nid': app.nid, 'phone': app.phone, 'email': app.email or '',
                'date_of_birth': app.date_of_birth,
                'center': app.chosen_center or app.routed_center,
                'gender': app.gender, 'education_qualification': app.education_qualification,
            },
        )
        self._log(request, f'Enrolled application {app.application_no}', app.id)
        return Response(ApplicationSerializer(app).data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        total = Circular.objects.count()
        published = Circular.objects.filter(status=Circular.Status.PUBLISHED).count()
        draft = Circular.objects.filter(status=Circular.Status.DRAFT).count()
        closed = Circular.objects.filter(status=Circular.Status.CLOSED).count()
        total_apps = Application.objects.count()
        active_apps = Application.objects.filter(status=Application.ApplicationStatus.PENDING).count()
        total_seats = Circular.objects.aggregate(s=Count('total_seats'))['s'] or 0
        return Response({
            'total_circulars': total,
            'published': published,
            'draft': draft,
            'closed': closed,
            'total_applications': total_apps,
            'active_applications': active_apps,
            'total_seats': total_seats,
        })

    @action(detail=True, methods=['get'])
    def print_application(self, request, pk=None):
        circular = self.get_object()
        app_id = request.query_params.get('application_id')
        if not app_id:
            return Response({'error': 'application_id required'}, status=400)
        try:
            app = Application.objects.select_related(
                'chosen_center', 'circular__course', 'user',
            ).get(id=app_id, circular=circular)
        except Application.DoesNotExist:
            return Response({'error': 'আবেদন পাওয়া যায়নি'}, status=404)

        html = render_to_string('applications/print_application.html', {
            'app': app,
            'circular': circular,
            'center': app.chosen_center,
            'course': circular.course,
        })
        pdf = HTML(string=html).write_pdf()
        filename = f'application_{app.application_no}.pdf'
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response

    @action(detail=True, methods=['get'])
    def print_selected_list(self, request, pk=None):
        circular = self.get_object()
        selected = Application.objects.filter(
            circular=circular,
            status=Application.ApplicationStatus.SELECTED,
        ).select_related('chosen_center').order_by('name_bn')

        html = render_to_string('applications/print_selected_list.html', {
            'circular': circular,
            'applications': selected,
            'course': circular.course,
        })
        pdf = HTML(string=html).write_pdf()
        filename = f'selected_list_{circular.public_url}.pdf'
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response

    @action(detail=False, methods=['get'], url_path='export-list')
    def export_list(self, request):
        qs = self.filter_queryset(self.get_queryset())
        ids = request.query_params.get('ids', '')
        if ids:
            id_list = [int(x) for x in ids.split(',') if x.isdigit()]
            if id_list:
                qs = qs.filter(id__in=id_list)
        fmt = request.query_params.get('file_format', 'xlsx')

        headers = [
            'শিরোনাম (বাংলা)', 'শিরোনাম (ইংরেজি)', 'সার্কুলার নম্বর',
            'কোর্স', 'মোট আসন', 'অবশিষ্ট আসন', 'ফি',
            'আবেদন শুরুর তারিখ', 'আবেদনের শেষ তারিখ',
            'প্রশিক্ষণ শুরুর তারিখ', 'প্রশিক্ষণ শেষের তারিখ',
            'অবস্থা',
        ]
        rows = []
        for c in qs:
            rows.append([
                c.title_bn, c.title_en, c.circular_no or '',
                c.course.code if c.course else '', c.total_seats, c.remaining_seats,
                str(c.fee) if c.fee else '',
                c.application_start_date.isoformat() if c.application_start_date else '',
                c.application_end_date.isoformat() if c.application_end_date else '',
                c.training_start_date.isoformat() if c.training_start_date else '',
                c.training_end_date.isoformat() if c.training_end_date else '',
                c.status,
            ])

        if fmt == 'xlsx':
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Circulars'
            ws.append(headers)
            for row in rows:
                ws.append(row)
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="circulars_{date.today().isoformat()}.xlsx"'
            return response
        else:
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            for row in rows:
                writer.writerow(row)
            response = HttpResponse(
                output.getvalue().encode('utf-8-sig'),
                content_type='text/csv',
            )
            response['Content-Disposition'] = f'attachment; filename="circulars_{date.today().isoformat()}.csv"'
            return response

    @action(detail=False, methods=['get'], url_path='download-template')
    def download_template(self, request):
        headers = [
            'শিরোনাম (বাংলা)', 'শিরোনাম (ইংরেজি)', 'সার্কুলার নম্বর',
            'কোর্স', 'মোট আসন', 'অবশিষ্ট আসন', 'ফি',
            'আবেদন শুরুর তারিখ', 'আবেদনের শেষ তারিখ',
            'প্রশিক্ষণ শুরুর তারিখ', 'প্রশিক্ষণ শেষের তারিখ',
            'অবস্থা',
        ]
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Circulars'
        ws.append(headers)
        ws.append(['উদাহরণ শিরোনাম বাংলা', 'উদাহরণ শিরোনাম English', 'BRTC/সার্কুলার/২০২৬/০১',
                    'C001', 50, 50, '5000.00',
                    '2026-01-01', '2026-01-31',
                    '2026-02-01', '2026-05-31',
                    'draft'])
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="circular_import_template.xlsx"'
        return response

    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def import_list(self, request):
        file = request.FILES.get('file')
        if not file:
            return Response({'error': 'ফাইল নির্বাচন করুন'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            rows_iter = iter(ws.iter_rows(values_only=True))
            header_row = [str(c).strip().lower() if c is not None else '' for c in next(rows_iter)]
        except Exception:
            file.seek(0)
            try:
                content = file.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(content))
                header_row = [h.strip().lower() for h in reader.fieldnames]
                rows_iter = reader
            except Exception:
                return Response({'error': 'ভুল ফাইল ফরম্যাট। Excel (.xlsx) বা CSV ফাইল আপলোড করুন।'}, status=400)

        bn_required = {'শিরোনাম (বাংলা)'}
        header_set = set(header_row)
        if not bn_required.issubset(header_set):
            return Response({
                'error': 'প্রয়োজনীয় কলাম নেই। হেডারে "শিরোনাম (বাংলা)" থাকা আবশ্যক।',
                'detected_headers': header_row,
            }, status=400)

        field_map = {
            'শিরোনাম (বাংলา)': 'title_bn', 'title_bn': 'title_bn',
            'শিরোনাম (ইংরেজি)': 'title_en', 'title_en': 'title_en',
            'সার্কুলার নম্বর': 'circular_no', 'circular_no': 'circular_no',
            'কোর্স': 'course_code',
            'মোট আসন': 'total_seats', 'total_seats': 'total_seats',
            'অবশিষ্ট আসন': 'remaining_seats', 'remaining_seats': 'remaining_seats',
            'ফি': 'fee', 'fee': 'fee',
            'আবেদন শুরুর তারিখ': 'application_start_date',
            'আবেদনের শেষ তারিখ': 'application_end_date',
            'প্রশিক্ষণ শুরুর তারিখ': 'training_start_date',
            'প্রশিক্ষণ শেষের তারিখ': 'training_end_date',
            'অবস্থা': 'status', 'status': 'status',
        }

        results = {'updated': 0, 'errors': []}
        for row_idx, row in enumerate(rows_iter, start=2):
            try:
                if isinstance(row, dict):
                    raw = row
                else:
                    raw = dict(zip(header_row, [str(c).strip() if c is not None else '' for c in row]))

                data = {}
                for k, v in raw.items():
                    mapped = field_map.get(k.strip().lower(), k.strip().lower())
                    data[mapped] = v.strip() if v else ''

                title_bn = data.get('title_bn', '').strip()
                if not title_bn:
                    results['errors'].append(f'সারি {row_idx}: শিরোনাম (বাংলা) খালি')
                    continue

                try:
                    circular = Circular.objects.get(title_bn=title_bn)
                except Circular.DoesNotExist:
                    results['errors'].append(f'সারি {row_idx}: "{title_bn}" পাওয়া যায়নি')
                    continue
                except Circular.MultipleObjectsReturned:
                    results['errors'].append(f'সারি {row_idx}: "{title_bn}" একাধিক সার্কুলার আছে')
                    continue

                if data.get('title_en'):
                    circular.title_en = data['title_en']
                if data.get('circular_no'):
                    circular.circular_no = data['circular_no']
                if data.get('total_seats'):
                    try:
                        circular.total_seats = int(data['total_seats'])
                    except ValueError:
                        pass
                if data.get('remaining_seats'):
                    try:
                        circular.remaining_seats = int(data['remaining_seats'])
                    except ValueError:
                        pass
                if data.get('fee'):
                    try:
                        circular.fee = float(data['fee'])
                    except ValueError:
                        pass
                for date_field in ('application_start_date', 'application_end_date',
                                    'training_start_date', 'training_end_date'):
                    if data.get(date_field):
                        parsed = parse_date(data[date_field])
                        if parsed:
                            setattr(circular, date_field, parsed)
                if data.get('status') in dict(Circular.Status.choices):
                    circular.status = data['status']

                circular.save()
                results['updated'] += 1

            except Exception as e:
                results['errors'].append(f'সারি {row_idx}: {str(e)}')

        return Response(results)

    @action(detail=True, methods=['get'], permission_classes=[permissions.AllowAny])
    def print_circular(self, request, pk=None):
        if not request.user.is_authenticated:
            token = request.query_params.get('token')
            if token:
                from rest_framework_simplejwt.tokens import AccessToken
                try:
                    user_id = AccessToken(token).payload.get('user_id')
                    from django.contrib.auth import get_user_model
                    request.user = get_user_model().objects.get(id=user_id)
                except Exception:
                    pass
            if not request.user.is_authenticated:
                return Response({'detail': 'Authentication credentials were not provided.'},
                                status=status.HTTP_401_UNAUTHORIZED)
        circular = self.get_object()
        centers = []
        if not circular.all_centers:
            centers = circular.eligible_centers.all()
        font_path = settings.BASE_DIR / 'static' / 'fonts' / 'NikoshBAN.ttf'
        html = render_to_string('circulars/print_circular.html', {
            'circular': circular,
            'centers': centers,
            'course': circular.course,
            'font_path': font_path,
        })
        pdf = HTML(string=html).write_pdf()
        filename = f'circular_{circular.public_url or circular.id}.pdf'
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'inline; filename="{filename}"'
        return response

from datetime import datetime

from django.db.models import Q
from django.utils import timezone

from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response

from django_filters.rest_framework import DjangoFilterBackend, DateFilter, CharFilter
from django_filters import FilterSet

from apps.circulars.models import Circular
from .models import Application
from .serializers import ApplicationExportSerializer
from .serializers_center import (
    ApplicationCenterListSerializer,
    ApplicationCenterDetailSerializer,
    ApplicationReviewSerializer,
    ApplicationBulkReviewSerializer,
)


class IsCenterAdminOrHeadOffice(permissions.BasePermission):
    def has_permission(self, request, view):
        if not request.user.is_authenticated:
            return False
        return (
            request.user.user_type in ('center_admin', 'head_office')
            or request.user.is_superuser
        )


class CenterApplicationFilter(FilterSet):
    date_from = DateFilter(field_name='applied_at', lookup_expr='gte')
    date_to = DateFilter(field_name='applied_at', lookup_expr='lte')
    name = CharFilter(method='filter_name')

    class Meta:
        model = Application
        fields = {
            'circular': ['exact'],
            'status': ['exact'],
            'nid': ['exact'],
            'phone': ['exact'],
        }

    def filter_name(self, queryset, name, value):
        return queryset.filter(
            Q(name_bn__icontains=value) | Q(name_en__icontains=value)
        )


class ApplicationCenterViewSet(viewsets.ReadOnlyModelViewSet):
    permission_classes = [IsCenterAdminOrHeadOffice]
    filter_backends = [DjangoFilterBackend]
    filterset_class = CenterApplicationFilter
    ordering = ('-applied_at',)

    def get_serializer_class(self):
        if self.action == 'list':
            return ApplicationCenterListSerializer
        return ApplicationCenterDetailSerializer

    def get_queryset(self):
        qs = Application.objects.select_related(
            'circular', 'circular__center', 'circular__course', 'reviewed_by'
        ).all()

        user = self.request.user
        if user.user_type == 'center_admin' and user.center:
            center_circulars = Circular.objects.filter(center=user.center).values_list('id', flat=True)
            qs = qs.filter(circular_id__in=center_circulars)
        elif user.user_type == 'center_admin' and not user.center:
            return Application.objects.none()

        return qs

    @action(detail=True, methods=['post'])
    def review(self, request, pk=None):
        application = self.get_object()
        serializer = ApplicationReviewSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        new_status = serializer.validated_data['status']
        remarks = serializer.validated_data.get('remarks', '')

        if application.status == new_status:
            return Response(
                {'error': f'আবেদনটি ইতিমধ্যে {application.get_status_display()} অবস্থায় আছে'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        application.status = new_status
        application.reviewed_by = request.user
        application.reviewed_at = timezone.now()
        if remarks:
            application.remarks = remarks
        application.save()

        return Response(ApplicationCenterDetailSerializer(application).data)

    @action(detail=False, methods=['post'])
    def bulk_review(self, request):
        serializer = ApplicationBulkReviewSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        ids = serializer.validated_data['ids']
        new_status = serializer.validated_data['status']
        remarks = serializer.validated_data.get('remarks', '')

        apps = self.get_queryset().filter(id__in=ids)
        updated = []
        errors = []

        for app in apps:
            if app.status == new_status:
                errors.append({'id': app.id, 'error': f'ইতিমধ্যে {app.get_status_display()} অবস্থায় আছে'})
                continue
            app.status = new_status
            app.reviewed_by = request.user
            app.reviewed_at = timezone.now()
            if remarks:
                app.remarks = remarks
            app.save()
            updated.append(app.id)

        return Response({
            'updated_count': len(updated),
            'updated_ids': updated,
            'error_count': len(errors),
            'errors': errors,
        })

    @action(detail=False, methods=['get'])
    def circulars(self, request):
        user = request.user
        qs = Circular.objects.filter(status=Circular.Status.PUBLISHED)
        if user.user_type == 'center_admin' and user.center:
            qs = qs.filter(center=user.center)
        data = [
            {'id': c.id, 'title_bn': c.title_bn, 'code': c.center.code}
            for c in qs.only('id', 'title_bn', 'center__code').select_related('center')
        ]
        return Response(data)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        qs = self.get_queryset()
        return Response({
            'total': qs.count(),
            'pending': qs.filter(status=Application.ApplicationStatus.PENDING).count(),
            'selected': qs.filter(status=Application.ApplicationStatus.SELECTED).count(),
            'rejected': qs.filter(status=Application.ApplicationStatus.REJECTED).count(),
            'waitlisted': qs.filter(status=Application.ApplicationStatus.WAITLISTED).count(),
        })

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        from django.http import HttpResponse
        import io
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
        except ImportError:
            return Response(
                {'error': 'openpyxl ইনস্টল করা নেই'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        queryset = self.filter_queryset(self.get_queryset())
        serializer = ApplicationExportSerializer(queryset, many=True)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Applications'

        headers = [
            'আবেদন নম্বর', 'নাম (বাংলায়)', 'নাম (ইংরেজিতে)',
            'পিতার নাম', 'মাতার নাম', 'এনআইডি', 'মোবাইল',
            'জন্ম তারিখ', 'ঠিকানা',
            'সার্কুলার', 'অবস্থা', 'আবেদনের তারিখ',
        ]

        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        for row_num, item in enumerate(serializer.data, 2):
            vals = [
                item.get('application_no', ''),
                item.get('name_bn', ''),
                item.get('name_en', ''),
                item.get('father_name_bn', ''),
                item.get('mother_name_bn', ''),
                item.get('nid', ''),
                item.get('phone', ''),
                item.get('date_of_birth', ''),
                item.get('present_address', ''),
                item.get('circular_title', ''),
                item.get('status_display', ''),
                item.get('applied_at', ''),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_num, column=col, value=val)
                cell.border = thin_border

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'applications_{datetime.now():%Y%m%d_%H%M%S}.xlsx'
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        from django.http import HttpResponse
        from django.template.loader import render_to_string

        queryset = self.filter_queryset(self.get_queryset())
        serializer = ApplicationExportSerializer(queryset, many=True)

        html = render_to_string('applications/export_pdf.html', {
            'applications': serializer.data,
            'generated_at': datetime.now().strftime('%d-%m-%Y %H:%M:%S'),
        })

        try:
            from weasyprint import HTML
        except ImportError:
            return Response({'error': 'WeasyPrint ইনস্টল করা নেই'}, status=500)

        pdf = HTML(string=html).write_pdf()

        filename = f'applications_{datetime.now():%Y%m%d_%H%M%S}.pdf'
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

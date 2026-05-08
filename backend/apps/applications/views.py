import io
import tempfile
from datetime import datetime

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.db.models import Q
from django.utils import timezone

from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from django_filters.rest_framework import DjangoFilterBackend, DateFilter, CharFilter
from django_filters import FilterSet

from .models import Application
from .serializers import (
    ApplicationListSerializer,
    ApplicationDetailSerializer,
    ApplicationWriteSerializer,
    ApplicationStatusUpdateSerializer,
    ApplicationExportSerializer,
)


class ApplicationFilter(FilterSet):
    date_from = DateFilter(field_name='applied_at', lookup_expr='gte')
    date_to = DateFilter(field_name='applied_at', lookup_expr='lte')
    name = CharFilter(method='filter_name')

    class Meta:
        model = Application
        fields = {
            'circular': ['exact'],
            'status': ['exact'],
            'nid': ['exact'],
            'phone': ['exact'],
        }

    def filter_name(self, queryset, name, value):
        return queryset.filter(
            Q(name_bn__icontains=value) | Q(name_en__icontains=value)
        )


class ApplicationViewSet(viewsets.ModelViewSet):
    queryset = Application.objects.select_related(
        'circular', 'circular__center', 'circular__course', 'reviewed_by'
    )
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = ApplicationFilter
    search_fields = ('application_no', 'name_bn', 'name_en', 'nid', 'phone')
    ordering_fields = ('applied_at', 'name_bn', 'status')
    ordering = ('-applied_at',)

    def get_serializer_class(self):
        if self.action == 'list':
            return ApplicationListSerializer
        elif self.action in ('create', 'update', 'partial_update'):
            return ApplicationWriteSerializer
        elif self.action in ('export_excel', 'export_pdf'):
            return ApplicationExportSerializer
        return ApplicationDetailSerializer

    @action(detail=True, methods=['patch'])
    def update_status(self, request, pk=None):
        application = self.get_object()
        serializer = ApplicationStatusUpdateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        application.status = serializer.validated_data['status']
        application.reviewed_by = request.user
        application.reviewed_at = timezone.now()
        if serializer.validated_data.get('remarks'):
            application.remarks = serializer.validated_data['remarks']
        application.save()

        return Response(ApplicationDetailSerializer(application).data)

    @action(detail=False, methods=['get'])
    def by_circular(self, request):
        circular_id = request.query_params.get('circular_id')
        if not circular_id:
            return Response(
                {'error': 'circular_id প্যারামিটার প্রয়োজন'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        queryset = self.filter_queryset(self.get_queryset()).filter(
            circular_id=circular_id
        )
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = ApplicationListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        serializer = ApplicationListSerializer(queryset, many=True)
        return Response(serializer.data)

    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = ApplicationExportSerializer(queryset, many=True)

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, Border, Side
        except ImportError:
            return Response(
                {'error': 'openpyxl ইনস্টল করা নেই। pip install openpyxl চালান।'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Applications'

        headers = [
            'আবেদন নম্বর', 'নাম (বাংলায়)', 'নাম (ইংরেজিতে)',
            'পিতার নাম', 'মাতার নাম', 'স্বামী/স্ত্রীর নাম',
            'জন্ম তারিখ', 'এনআইডি', 'মোবাইল', 'বিকল্প মোবাইল', 'ইমেইল',
            'বর্তমান ঠিকানা', 'স্থায়ী ঠিকানা',
            'শিক্ষাগত যোগ্যতা', 'পেশা',
            'সার্কুলার', 'কেন্দ্র', 'কোর্স',
            'অবস্থা', 'আবেদনের তারিখ', 'পর্যালোচনার তারিখ', 'মন্তব্য',
        ]

        header_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'),
        )

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        for row_num, item in enumerate(serializer.data, 2):
            ws.cell(row=row_num, column=1, value=item.get('application_no', ''))
            ws.cell(row=row_num, column=2, value=item.get('name_bn', ''))
            ws.cell(row=row_num, column=3, value=item.get('name_en', ''))
            ws.cell(row=row_num, column=4, value=item.get('father_name_bn', ''))
            ws.cell(row=row_num, column=5, value=item.get('mother_name_bn', ''))
            ws.cell(row=row_num, column=6, value=item.get('spouse_name_bn', ''))
            ws.cell(row=row_num, column=7, value=item.get('date_of_birth', ''))
            ws.cell(row=row_num, column=8, value=item.get('nid', ''))
            ws.cell(row=row_num, column=9, value=item.get('phone', ''))
            ws.cell(row=row_num, column=10, value=item.get('alternate_phone', ''))
            ws.cell(row=row_num, column=11, value=item.get('email', ''))
            ws.cell(row=row_num, column=12, value=item.get('present_address', ''))
            ws.cell(row=row_num, column=13, value=item.get('permanent_address', ''))
            ws.cell(row=row_num, column=14, value=item.get('education_qualification', ''))
            ws.cell(row=row_num, column=15, value=item.get('profession', ''))
            ws.cell(row=row_num, column=16, value=item.get('circular_title', ''))
            ws.cell(row=row_num, column=17, value=item.get('center_name', ''))
            ws.cell(row=row_num, column=18, value=item.get('course_name', ''))
            ws.cell(row=row_num, column=19, value=item.get('status_display', ''))
            ws.cell(row=row_num, column=20, value=item.get('applied_at', ''))
            ws.cell(row=row_num, column=21, value=item.get('reviewed_at', ''))
            ws.cell(row=row_num, column=22, value=item.get('remarks', ''))
            for col in range(1, 23):
                ws.cell(row=row_num, column=col).border = thin_border

        for col in range(1, 23):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f'applications_{datetime.now():%Y%m%d_%H%M%S}.xlsx'
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=['get'])
    def export_pdf(self, request):
        queryset = self.filter_queryset(self.get_queryset())
        serializer = ApplicationExportSerializer(queryset, many=True)

        html = render_to_string('applications/export_pdf.html', {
            'applications': serializer.data,
            'generated_at': datetime.now().strftime('%d-%m-%Y %H:%M:%S'),
        })

        try:
            from weasyprint import HTML
        except ImportError:
            return Response(
                {'error': 'WeasyPrint ইনস্টল করা নেই।'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        pdf = HTML(string=html).write_pdf()

        filename = f'applications_{datetime.now():%Y%m%d_%H%M%S}.pdf'
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
